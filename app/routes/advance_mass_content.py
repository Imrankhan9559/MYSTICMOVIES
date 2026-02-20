import asyncio
import csv
import io
import logging
import os
import re
from datetime import datetime, timedelta
from typing import Any

from beanie.operators import In
from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from app.core.config import settings
from app.db.models import ContentItem, FileSystemItem, MassContentState, User
from app.routes.admin import _admin_context_base, _build_title_regex, _is_admin, _publish_items
from app.routes.content import _parse_name, _tmdb_details, _tmdb_get, _tmdb_search
from app.routes.dashboard import _cast_ids, get_current_user
from app.utils.file_utils import format_size

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")
logger = logging.getLogger(__name__)

_MASS_WS_CLIENTS: set[WebSocket] = set()
_MASS_TASKS: dict[str, asyncio.Task] = {}
_MASS_UPLOAD_TASKS: dict[str, asyncio.Task] = {}
_MASS_LOCKS: dict[str, asyncio.Lock] = {}
_MASS_BROADCAST_LOCK = asyncio.Lock()
_MASS_PROCESS_WORKERS = max(1, min(int(os.getenv("MASS_PROCESS_WORKERS", "3")), 8))
_MASS_UPLOAD_WORKERS = max(1, min(int(os.getenv("MASS_UPLOAD_WORKERS", "3")), 8))
_MASS_PROCESS_SEMAPHORE = asyncio.Semaphore(_MASS_PROCESS_WORKERS)
_MASS_UPLOAD_SEMAPHORE = asyncio.Semaphore(_MASS_UPLOAD_WORKERS)
_MASS_SNAPSHOT_CACHE: dict[str, Any] = {"ts": 0.0, "data": None, "dirty": True}
_MASS_SNAPSHOT_MIN_INTERVAL_SEC = 0.8
_MASS_SNAPSHOT_IDLE_CACHE_TTL_SEC = max(2.0, float(os.getenv("MASS_SNAPSHOT_IDLE_CACHE_TTL_SEC", "300")))
_STORAGE_POOL_CACHE: dict[str, Any] = {"rows": [], "expires_at": datetime.min}
_STORAGE_POOL_TTL_SECONDS = 45
_IMPORT_CHUNK_SIZE = 30
_MASS_IMPORT_LOCK = asyncio.Lock()
_MASS_IMPORT_TASK: asyncio.Task | None = None
_MASS_BG_DEDUPE_TASK: asyncio.Task | None = None
_MASS_BG_UPLOAD_RECOVERY_TASK: asyncio.Task | None = None
_MASS_IMPORT_QUEUE: list[dict[str, str]] = []
_MASS_MAX_PENDING_PROCESS_TASKS = 120
_MASS_SNAPSHOT_ROW_LIMIT = max(120, min(int(os.getenv("MASS_SNAPSHOT_ROW_LIMIT", "320")), 3000))
_MASS_DEDUPE_SCAN_LIMIT = max(500, min(int(os.getenv("MASS_DEDUPE_SCAN_LIMIT", "12000")), 50000))
_MASS_DEDUPE_MAX_RUNTIME_SEC = max(0.5, float(os.getenv("MASS_DEDUPE_MAX_RUNTIME_SEC", "3.0")))
_MASS_EXPORT_ROW_LIMIT = max(500, min(int(os.getenv("MASS_EXPORT_ROW_LIMIT", "20000")), 50000))
_MASS_DEDUPE_COOLDOWN_SEC = 180
_MASS_DEDUPE_LAST_RUN_TS = 0.0
_MASS_UPLOAD_STALE_SEC = max(60, int(os.getenv("MASS_UPLOAD_STALE_SEC", "900")))
_MASS_UPLOAD_RECOVERY_COOLDOWN_SEC = max(10, int(os.getenv("MASS_UPLOAD_RECOVERY_COOLDOWN_SEC", "60")))
_MASS_UPLOAD_RECOVERY_LIMIT = max(10, min(int(os.getenv("MASS_UPLOAD_RECOVERY_LIMIT", "200")), 2000))
_MASS_UPLOAD_RECOVERY_LAST_RUN_TS = 0.0
_MASS_IMPORT_STATUS: dict[str, Any] = {
    "running": False,
    "queued": 0,
    "processed": 0,
    "created": 0,
    "failed": 0,
    "chunk_size": _IMPORT_CHUNK_SIZE,
    "started_at": "",
    "updated_at": "",
    "message": "Idle",
}


def _collection_for(model_cls):
    for attr in ("get_motor_collection", "get_pymongo_collection", "get_collection"):
        getter = getattr(model_cls, attr, None)
        if callable(getter):
            return getter()
    raise AttributeError(f"No collection getter found for {model_cls!r}")


class _MassRowProxy:
    __slots__ = ("_doc",)

    def __init__(self, doc: dict[str, Any]):
        self._doc = doc or {}

    def __getattr__(self, name: str):
        if name == "id":
            return self._doc.get("id") or self._doc.get("_id")
        return self._doc.get(name)


_SERIES_SE_RE = re.compile(r"[Ss](\d{1,2})[\s._-]*[Ee](\d{1,3})")
_SERIES_SE_ALT_RE = re.compile(r"\b(\d{1,2})x(\d{1,3})\b", re.I)
_SERIES_WORD_RE = re.compile(r"\bSeason\s*(\d{1,2})\b.*?\bEpisode\s*(\d{1,3})\b", re.I)
_YEAR_TOKEN_RE = re.compile(r"(?<!\d)((?:19|20)\d{2})(?!\d)")
_TMDB_SEASON_EPISODE_NOISE_RE = re.compile(
    r"\b(?:s\d{1,2}[\s._-]*e\d{1,3}|s\d{1,2}|e\d{1,3}|season\s*\d{1,2}|episode\s*\d{1,3}|ep\s*\d{1,3}|\d{1,2}x\d{1,3})\b",
    re.I,
)
_TMDB_QUALITY_NOISE_RE = re.compile(r"\b(?:2160|1440|1080|720|480|380|360)p?\b|\b(?:4k|uhd|fhd|hd)\b", re.I)
_TMDB_FILE_NOISE_RE = re.compile(
    r"\b(?:webrip|web[-\s]?dl|webdl|bluray|brrip|hdrip|dvdrip|cam|hdcam|hdts|hdtc|proper|repack|internal|"
    r"remux|extended|uncut|multi|audio|dual|dub|dubbed|subs?|esub|x264|x265|h264|h265|hevc|aac|ac3|ddp|"
    r"atmos|10bit|8bit|mkv|mp4|avi|webm|m4v|pack|complete)\b",
    re.I,
)

_TITLE_STOP_TOKENS = {
    "the", "a", "an", "and", "or", "to", "of", "for", "with", "in", "on", "at", "by", "from",
    "this", "that", "movie", "film", "series", "show", "web", "part", "vol", "volume",
    "chapter", "season", "episode", "ep",
}
_FILE_NOISE_TOKENS = {
    "webrip", "web", "webdl", "webd", "bluray", "brrip", "hdrip", "dvdrip", "cam", "ts",
    "proper", "repack", "internal", "remux", "extended", "uncut", "multi", "audio", "dub",
    "subs", "sub", "english", "hindi", "tamil", "telugu", "malayalam", "bengali", "korean",
    "japanese", "nf", "amzn", "dsnp", "hdr", "hdr10", "dv", "x264", "x265", "h264", "h265",
    "hevc", "aac", "ac3", "ddp", "atmos", "bit", "mkv", "mp4", "avi", "webm", "m4v",
}


def _clean_match_tokens(text: str, *, for_file_name: bool = False) -> set[str]:
    out: set[str] = set()
    for token in _tokens(text):
        tok = (token or "").strip().lower()
        if not tok:
            continue
        if tok in _TITLE_STOP_TOKENS:
            continue
        # Ignore common season/episode marker tokens on both sides
        # (e.g., season2, episode08, s02, e08, 2x08), so noisy titles still match.
        if re.fullmatch(r"s\d{1,2}", tok):
            continue
        if re.fullmatch(r"e\d{1,3}", tok):
            continue
        if re.fullmatch(r"season\d{1,2}", tok):
            continue
        if re.fullmatch(r"episode\d{1,3}", tok):
            continue
        if re.fullmatch(r"ep\d{1,3}", tok):
            continue
        if re.fullmatch(r"\d{1,2}x\d{1,3}", tok):
            continue
        # Ignore quality/format marker tokens on both sides so "title + 480p"
        # can still match files even if candidate-side token filtering removes quality tokens.
        if re.fullmatch(r"\d{3,4}p", tok):
            continue
        if tok in {"4k", "2k", "8k", "uhd", "fhd", "hd"}:
            continue
        if for_file_name:
            if tok in _FILE_NOISE_TOKENS:
                continue
            if re.fullmatch(r"[xh]\d{3,4}", tok):
                continue
            if re.fullmatch(r"\d{1,2}bit", tok):
                continue
        # Ignore release-year tokens for matching.
        if re.fullmatch(r"(19|20)\d{2}", tok):
            continue
        # Remove tiny noise tokens while keeping meaningful numerics.
        if len(tok) == 1 and not tok.isdigit():
            continue
        out.add(tok)
    return out


def _title_signature(text: str) -> str:
    toks = sorted(_clean_match_tokens(text, for_file_name=False))
    if not toks:
        return _norm_title(text)
    return "-".join(toks)


def _norm_title(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (text or "").lower())


def _tokens(text: str) -> list[str]:
    return [x for x in re.split(r"[^a-z0-9]+", (text or "").lower()) if x]


def _clean_tmdb_query_base(text: str, *, drop_year: bool) -> str:
    raw = " ".join((text or "").replace("_", " ").replace(".", " ").split()).strip()
    if not raw:
        return ""
    clean = _TMDB_SEASON_EPISODE_NOISE_RE.sub(" ", raw)
    clean = _TMDB_QUALITY_NOISE_RE.sub(" ", clean)
    clean = _TMDB_FILE_NOISE_RE.sub(" ", clean)
    if drop_year:
        clean = _YEAR_TOKEN_RE.sub(" ", clean)
    clean = re.sub(r"[\[\]\(\)\{\}|]+", " ", clean)
    clean = re.sub(r"\s+", " ", clean).strip(" -_:,")
    return clean


def _ordered_tmdb_tokens(text: str, *, drop_year: bool = True) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for token in _tokens(text):
        tok = (token or "").strip().lower()
        if not tok:
            continue
        if tok in _TITLE_STOP_TOKENS or tok in _FILE_NOISE_TOKENS:
            continue
        if re.fullmatch(r"s\d{1,2}", tok) or re.fullmatch(r"e\d{1,3}", tok):
            continue
        if re.fullmatch(r"\d{1,2}x\d{1,3}", tok):
            continue
        if re.fullmatch(r"\d{3,4}p", tok):
            continue
        if re.fullmatch(r"[xh]\d{3,4}", tok):
            continue
        if re.fullmatch(r"\d{1,2}bit", tok):
            continue
        if drop_year and re.fullmatch(r"(19|20)\d{2}", tok):
            continue
        if len(tok) == 1 and not tok.isdigit():
            continue
        if tok in seen:
            continue
        seen.add(tok)
        out.append(tok)
    return out


def _tmdb_query_variants(raw_title: str) -> list[str]:
    clean = " ".join((raw_title or "").split()).strip()
    if not clean:
        return []

    variants: list[str] = [clean]

    parsed = _parse_name(clean)
    parsed_title = " ".join((parsed.get("title") or "").split()).strip()
    if parsed_title:
        variants.append(parsed_title)

    stripped_keep_year = _clean_tmdb_query_base(clean, drop_year=False)
    stripped_no_year = _clean_tmdb_query_base(clean, drop_year=True)
    if stripped_keep_year:
        variants.append(stripped_keep_year)
    if stripped_no_year:
        variants.append(stripped_no_year)

    for base in [stripped_no_year, stripped_keep_year, parsed_title, clean]:
        if not base:
            continue
        ordered_tokens = _ordered_tmdb_tokens(base, drop_year=True)
        if ordered_tokens:
            variants.append(" ".join(ordered_tokens))

    expanded: list[str] = []
    for v in variants:
        expanded.extend(_expand_title_variants(v))

    seen: set[str] = set()
    out: list[str] = []
    for value in expanded:
        text = " ".join((value or "").split()).strip()
        if not text:
            continue
        key = _norm_title(text)
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(text)
    return out[:8]


def _tmdb_result_title(result: dict, is_series: bool) -> str:
    if is_series:
        return " ".join(str(result.get("name") or result.get("title") or "").split()).strip()
    return " ".join(str(result.get("title") or result.get("name") or "").split()).strip()


def _tmdb_result_year(result: dict, is_series: bool) -> int | None:
    date_value = (
        result.get("first_air_date")
        if is_series
        else result.get("release_date")
    ) or result.get("release_date") or result.get("first_air_date") or ""
    return _extract_primary_year(str(date_value))


def _tmdb_score_result(
    result: dict,
    intent_tokens: set[str],
    intent_norm: str,
    intent_year: int | None,
    is_series: bool,
) -> float:
    title = _tmdb_result_title(result, is_series)
    if not title:
        return -999.0

    result_norm = _norm_title(title)
    result_tokens = _clean_match_tokens(title, for_file_name=False)
    if not result_tokens:
        return -998.0

    overlap = len(intent_tokens & result_tokens) if intent_tokens else 0
    if intent_tokens and overlap == 0 and intent_norm != result_norm:
        return -50.0

    coverage = (overlap / len(intent_tokens)) if intent_tokens else 0.0
    precision = (overlap / len(result_tokens)) if result_tokens else 0.0
    exact = 1.0 if intent_norm and result_norm == intent_norm else 0.0
    contains = 1.0 if intent_norm and (intent_norm in result_norm or result_norm in intent_norm) else 0.0

    year_bonus = 0.0
    result_year = _tmdb_result_year(result, is_series)
    if intent_year and result_year:
        year_bonus = 8.0 if int(intent_year) == int(result_year) else -2.0

    try:
        popularity_bonus = min(float(result.get("popularity") or 0.0), 100.0) / 25.0
    except Exception:
        popularity_bonus = 0.0

    return (
        (exact * 120.0)
        + (contains * 25.0)
        + (coverage * 70.0)
        + (precision * 20.0)
        + (overlap * 8.0)
        + year_bonus
        + popularity_bonus
    )


def _pick_best_tmdb_result(
    results: list[dict],
    query_variants: list[str],
    year: str,
    is_series: bool,
) -> tuple[dict | None, float]:
    if not results:
        return None, -999.0

    intents: list[tuple[set[str], str]] = []
    for query in query_variants:
        tokens = _clean_match_tokens(query, for_file_name=False)
        if not tokens:
            continue
        intents.append((tokens, _norm_title(query)))
    if not intents:
        return None, -998.0

    intent_year = _extract_primary_year(str(year or ""))
    best: dict | None = None
    best_score = -999.0

    for row in results:
        row_score = max(
            _tmdb_score_result(row, tokens, norm, intent_year, is_series)
            for tokens, norm in intents
        )
        if row_score > best_score:
            best_score = row_score
            best = row

    # Guardrail: avoid random weak matches.
    if best is None or best_score < 35.0:
        return None, best_score
    return best, best_score


async def _tmdb_search_best_for_type(raw_title: str, year: str, is_series: bool) -> dict[str, Any]:
    query_variants = _tmdb_query_variants(raw_title)
    if not query_variants:
        return {"pick": None, "results": [], "is_series": is_series, "score": -999.0}

    year_value = str(_extract_primary_year(str(year or "")) or "").strip()
    attempts: list[tuple[str, str]] = []
    for query in query_variants:
        if year_value:
            attempts.append((query, year_value))
        attempts.append((query, ""))

    seen_attempts: set[tuple[str, str]] = set()
    unique_attempts: list[tuple[str, str]] = []
    for query, year_hint in attempts:
        key = (_norm_title(query), year_hint)
        if not key[0] or key in seen_attempts:
            continue
        seen_attempts.add(key)
        unique_attempts.append((query, year_hint))

    seen_result_ids: set[str] = set()
    merged_results: list[dict] = []

    for query, year_hint in unique_attempts[:10]:
        search = await _tmdb_search(query, year_hint, is_series)
        rows = (search or {}).get("results") or []
        for result in rows:
            rid = result.get("id")
            if not rid:
                continue
            key = str(rid)
            if key in seen_result_ids:
                continue
            seen_result_ids.add(key)
            merged_results.append(result)

        pick, score = _pick_best_tmdb_result(merged_results, query_variants, year, is_series)
        if pick and score >= 95.0:
            return {"pick": pick, "results": merged_results, "is_series": is_series, "score": score}

    pick, score = _pick_best_tmdb_result(merged_results, query_variants, year, is_series)
    return {"pick": pick, "results": merged_results, "is_series": is_series, "score": score}


async def _tmdb_search_best(raw_title: str, year: str, is_series: bool) -> dict[str, Any]:
    primary = await _tmdb_search_best_for_type(raw_title, year, is_series)
    primary_pick = primary.get("pick")
    primary_score = float(primary.get("score") or -999.0)

    # Fast path: keep current inferred type when confidence is already strong.
    if primary_pick and primary_score >= 70.0:
        return primary

    alternate = await _tmdb_search_best_for_type(raw_title, year, not is_series)
    alternate_pick = alternate.get("pick")
    alternate_score = float(alternate.get("score") or -999.0)

    if primary_pick and alternate_pick:
        # Keep inferred type unless alternate is significantly stronger.
        return alternate if alternate_score >= (primary_score + 15.0) else primary
    if primary_pick:
        return primary
    if alternate_pick:
        return alternate
    return primary if primary_score >= alternate_score else alternate


def _to_title_marker(title: str) -> str:
    clean = " ".join((title or "").split()).strip()
    return f"title:{clean}" if clean else ""


def _to_source_marker(source_note: str) -> str:
    clean = (source_note or "").strip().lower()
    return f"source:{clean}" if clean else ""


def _extract_title_markers(values: list[str] | None) -> list[str]:
    out: list[str] = []
    for raw in values or []:
        text = str(raw or "").strip()
        if not text:
            continue
        lower = text.lower()
        if lower.startswith("title:"):
            t = text.split(":", 1)[1].strip()
            if t:
                out.append(t)
            continue
        if lower.startswith("source:"):
            continue
        if lower in {"manual", "csv_excel"}:
            continue
        # Backward compatibility: old rows may have saved plain titles directly.
        out.append(text)
    # Unique preserve order
    seen: set[str] = set()
    uniq: list[str] = []
    for t in out:
        key = _norm_title(t)
        if not key or key in seen:
            continue
        seen.add(key)
        uniq.append(t)
    return uniq


def _expand_title_variants(title: str) -> list[str]:
    clean = " ".join((title or "").split()).strip()
    if not clean:
        return []
    variants = [clean]
    parsed = _parse_name(clean)
    parsed_title = " ".join((parsed.get("title") or "").split()).strip()
    if parsed_title:
        variants.append(parsed_title)
    # Remove common suffix qualifiers.
    for sep in [":", "-", "|", "("]:
        if sep in clean:
            part = clean.split(sep, 1)[0].strip()
            if part and len(_tokens(part)) >= 2:
                variants.append(part)
    # Compact form with stop words removed helps with noisy filenames.
    stop = {"the", "a", "an", "movie", "series", "season", "episode", "part", "vol", "volume"}
    token_list = [t for t in _tokens(clean) if t not in stop]
    if token_list:
        variants.append(" ".join(token_list))
    # Unique preserve order
    seen: set[str] = set()
    out: list[str] = []
    for value in variants:
        key = _norm_title(value)
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(value)
    return out


def _guess_type(raw_title: str, hint: str) -> str:
    value = (hint or "").strip().lower()
    if value in {"movie", "series"}:
        return value
    name = (raw_title or "").lower()
    if re.search(r"\b(series|web\s*series|tv|season|episode|s\d{1,2}[\s._-]*e\d{1,3}|s\d{1,2}\b.*\be\d{1,3})\b", name):
        return "series"
    return "movie"


def _quality_rank(label: str) -> int:
    q = (label or "").upper().replace(" ", "")
    order = {"4K": 6, "2160P": 6, "1440P": 5, "1080P": 4, "720P": 3, "480P": 2, "360P": 1, "HD": 0}
    return order.get(q, 0)


def _series_quality_from_name(name: str) -> str:
    lower = (name or "").lower()
    if "2160" in lower or "4k" in lower:
        return "4K"
    if "1080" in lower:
        return "1080P"
    if "720" in lower:
        return "720P"
    if "480" in lower:
        return "480P"
    if "360" in lower or "380" in lower:
        return "360P"
    return "HD"


def _movie_quality_from_size(size: int) -> str:
    total = int(size or 0)
    gb = 1024 * 1024 * 1024
    mb = 1024 * 1024
    if total > 2 * gb:
        return "1080P"
    if total >= 1 * gb:
        return "720P"
    if total >= 500 * mb:
        return "480P"
    return "360P"


def _movie_quality_from_name_or_size(name: str, size: int) -> str:
    # Movies are intentionally inferred from file size only.
    # Web-series still use filename-based quality detection.
    return _movie_quality_from_size(size)


def _normalize_quality_label(value: str) -> str:
    raw = (value or "").strip().upper().replace(" ", "")
    if not raw:
        return "HD"
    aliases = {
        "2160": "4K",
        "2160P": "4K",
        "4KP": "4K",
        "4K": "4K",
        "1080": "1080P",
        "1080P": "1080P",
        "720": "720P",
        "720P": "720P",
        "480": "480P",
        "480P": "480P",
        "360": "360P",
        "360P": "360P",
        "380": "360P",
        "380P": "360P",
        "HD": "HD",
    }
    return aliases.get(raw, raw)


def _series_season_episode(name: str) -> tuple[int | None, int | None]:
    raw = name or ""
    match = _SERIES_SE_RE.search(raw)
    if match:
        return int(match.group(1)), int(match.group(2))
    match = _SERIES_SE_ALT_RE.search(raw)
    if match:
        return int(match.group(1)), int(match.group(2))
    match = _SERIES_WORD_RE.search(raw)
    if match:
        return int(match.group(1)), int(match.group(2))
    return None, None


def _int_or_none(value: Any) -> int | None:
    try:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        val = int(text)
        return val if val > 0 else None
    except Exception:
        return None


def _extract_years_from_text(text: str) -> set[int]:
    years: set[int] = set()
    for match in _YEAR_TOKEN_RE.finditer(text or ""):
        try:
            year = int(match.group(1))
        except Exception:
            continue
        if 1900 <= year <= 2099:
            years.add(year)
    return years


def _extract_primary_year(text: str) -> int | None:
    years = _extract_years_from_text(text or "")
    if not years:
        return None
    return sorted(years)[0]


def _content_release_year(state: MassContentState) -> int | None:
    # Prefer TMDB release date year. Fallback to imported/manual year when absent.
    tmdb_year = _extract_primary_year(str(getattr(state, "release_date", "") or ""))
    if tmdb_year:
        return tmdb_year
    return _extract_primary_year(str(getattr(state, "year", "") or ""))


def _is_video_row(item: FileSystemItem) -> bool:
    if item.is_folder:
        return False
    mime = (item.mime_type or "").lower()
    if mime.startswith("video"):
        return True
    lower = (item.name or "").lower()
    return lower.endswith((".mkv", ".mp4", ".avi", ".mov", ".wmv", ".webm", ".m4v"))


def _source_label(item: FileSystemItem) -> str:
    source = (item.source or "").strip().lower()
    if source == "bot":
        return "Bot Listening"
    if source == "upload":
        return "Dashboard Upload"
    return "Telegram Storage"


def _source_upload_label(item: FileSystemItem) -> str:
    source = (item.source or "").strip().lower()
    if source == "bot":
        return "Uploaded from Bot Listener"
    if source == "upload":
        return "Uploaded from Dashboard"
    return "Uploaded from Telegram Storage"


def _title_match(target_titles: list[str], file_name: str, row_title: str = "", row_series_title: str = "") -> bool:
    # Important: match by actual filename tokens, not mutable metadata fields.
    # This prevents false matches like "Venom ... Dance" for "Ivy + Bean: Doomed to Dance".
    parsed = _parse_name(file_name or "")
    parsed_title = (parsed.get("title") or "").strip()
    candidate_tokens = _clean_match_tokens(file_name or "", for_file_name=True)
    candidate_tokens.update(_clean_match_tokens(parsed_title, for_file_name=True))
    if not candidate_tokens:
        return False

    target_sets: list[tuple[str, set[str]]] = []
    for target in target_titles:
        clean = " ".join((target or "").split()).strip()
        if not clean:
            continue
        toks = _clean_match_tokens(clean, for_file_name=False)
        if toks:
            target_sets.append((clean, toks))
    if not target_sets:
        return False

    max_len = max(len(toks) for _, toks in target_sets)
    filename_norm = _norm_title(file_name or "")
    parsed_norm = _norm_title(parsed_title)

    # Strict policy: all important words from the strongest title variant must be present.
    for raw, toks in target_sets:
        if len(toks) != max_len:
            continue
        if toks.issubset(candidate_tokens):
            return True
        # Single-token title fallback for very short names.
        if len(toks) == 1:
            token = next(iter(toks))
            if token in candidate_tokens:
                raw_norm = _norm_title(raw)
                if raw_norm and (raw_norm in filename_norm or (parsed_norm and raw_norm in parsed_norm)):
                    return True

    # Secondary fallback for short names only (2 tokens max), still strict subset.
    if max_len <= 2:
        for _, toks in target_sets:
            if toks and toks.issubset(candidate_tokens):
                return True
    return False


def _panel_label(panel: str) -> str:
    labels = {
        "processing": "Searching / Processing",
        "tmdb_not_found": "TMDB Not Found",
        "file_not_found": "File Not Found",
        "incomplete": "Incomplete Content",
        "complete": "Complete Content",
        "uploading": "Uploading",
        "uploaded": "Uploaded Content",
        "skipped": "Skipped Existing",
    }
    key = _normalize_mass_panel_key(panel, fallback="")
    return labels.get(key, panel)


def _series_quality_coverage(row: MassContentState) -> list[dict]:
    if (row.content_type or "").strip().lower() != "series":
        return []

    season_expected: dict[int, set[int]] = {}
    for season_row in list(getattr(row, "seasons", []) or []):
        season_no = _int_or_none(season_row.get("season"))
        if not season_no:
            continue
        season_expected.setdefault(season_no, set())
        eps: set[int] = set()
        for ep in list(season_row.get("episodes", []) or []):
            ep_no = _int_or_none(ep.get("episode"))
            if ep_no:
                eps.add(ep_no)
        if eps:
            season_expected[season_no].update(eps)

    quality_hits: dict[int, dict[str, set[int]]] = {}
    quality_missing: dict[int, dict[str, set[int]]] = {}

    def _ensure_quality_bucket(store: dict[int, dict[str, set[int]]], season_no: int, quality: str) -> set[int]:
        return store.setdefault(season_no, {}).setdefault(quality, set())

    for item in list(getattr(row, "matched_files", []) or []):
        season_no = _int_or_none(item.get("season"))
        episode_no = _int_or_none(item.get("episode"))
        if not season_no or not episode_no:
            continue
        quality = _normalize_quality_label(item.get("quality") or "HD")
        _ensure_quality_bucket(quality_hits, season_no, quality).add(episode_no)
        season_expected.setdefault(season_no, set()).add(episode_no)

    for item in list(getattr(row, "missing_items", []) or []):
        season_no = _int_or_none(item.get("season"))
        episode_no = _int_or_none(item.get("episode"))
        if not season_no:
            continue
        season_expected.setdefault(season_no, set())
        quality = _normalize_quality_label(item.get("quality") or "HD")
        _ensure_quality_bucket(quality_hits, season_no, quality)
        _ensure_quality_bucket(quality_missing, season_no, quality)
        if episode_no:
            season_expected.setdefault(season_no, set()).add(episode_no)
            _ensure_quality_bucket(quality_missing, season_no, quality).add(episode_no)

    # Fallback from live notes so coverage remains complete even when row payload is partial/stale.
    for note in list(getattr(row, "live_notes", []) or []):
        if not isinstance(note, dict):
            continue
        season_no = _int_or_none(note.get("season"))
        episode_no = _int_or_none(note.get("episode"))
        if not season_no:
            continue
        season_expected.setdefault(season_no, set())
        quality = _normalize_quality_label(note.get("quality") or "HD")
        state = str(note.get("state") or "").strip().lower()
        _ensure_quality_bucket(quality_hits, season_no, quality)
        _ensure_quality_bucket(quality_missing, season_no, quality)
        if episode_no:
            season_expected[season_no].add(episode_no)
            if state == "missing":
                _ensure_quality_bucket(quality_missing, season_no, quality).add(episode_no)
            else:
                _ensure_quality_bucket(quality_hits, season_no, quality).add(episode_no)

    if not season_expected:
        return []

    payload: list[dict] = []
    for season_no in sorted(season_expected.keys()):
        expected_eps = {int(x) for x in season_expected.get(season_no, set()) if int(x) > 0}
        expected_count = len(expected_eps)
        by_quality = quality_hits.get(season_no, {})
        by_missing = quality_missing.get(season_no, {})
        # Hide placeholder seasons that have no expected episode data and no matched/missing rows.
        has_episode_evidence = bool(expected_eps)
        if not has_episode_evidence:
            any_found = any(bool(v) for v in by_quality.values()) if by_quality else False
            any_missing = any(bool(v) for v in by_missing.values()) if by_missing else False
            if not any_found and not any_missing:
                continue
        # Always show base qualities even when not found.
        for base_q in ("1080P", "720P", "480P"):
            by_quality.setdefault(base_q, set())
            by_missing.setdefault(base_q, set())

        # Merge any quality that is only in missing list.
        for quality in by_missing.keys():
            by_quality.setdefault(quality, set())
        quality_rows: list[dict] = []
        for quality, covered in by_quality.items():
            found_set = {int(x) for x in (covered or set()) if int(x) > 0}
            missing_set = {int(x) for x in (by_missing.get(quality) or set()) if int(x) > 0}
            if expected_eps:
                missing_set = (expected_eps - found_set) | missing_set
                found_count = len(found_set & expected_eps)
                expected_count_quality = len(expected_eps)
            else:
                expected_pool = found_set | missing_set
                found_count = len(found_set)
                expected_count_quality = len(expected_pool)
            quality_rows.append(
                {
                    "quality": quality,
                    "found": found_count,
                    "expected": expected_count_quality if expected_count_quality >= 0 else expected_count,
                    "found_episodes": sorted(list(found_set)),
                    "missing_episodes": sorted(list(missing_set)),
                }
            )
        quality_rows.sort(
            key=lambda x: (-_quality_rank(x.get("quality") or ""), (x.get("quality") or "").upper())
        )
        payload.append(
            {
                "season": season_no,
                "expected": expected_count,
                "qualities": quality_rows,
            }
        )
    return payload


_MASS_PANEL_KEYS = {
    "processing",
    "tmdb_not_found",
    "file_not_found",
    "incomplete",
    "complete",
    "uploading",
    "uploaded",
    "skipped",
}


def _normalize_mass_panel_key(panel: str, fallback: str = "processing") -> str:
    raw = str(panel or "").strip().lower().replace("-", "_").replace(" ", "_")
    aliases = {
        "tmdbnotfound": "tmdb_not_found",
        "filenotfound": "file_not_found",
        "not_found_tmdb": "tmdb_not_found",
        "not_found_file": "file_not_found",
        "searching": "processing",
    }
    raw = aliases.get(raw, raw)
    if raw in _MASS_PANEL_KEYS:
        return raw
    return fallback


def _file_note_suffix(row: dict) -> str:
    name = (row.get("name") or "").strip()
    size_label = format_size(int(row.get("size") or 0))
    if name:
        return f" | {name} ({size_label})"
    return f" | {size_label}"


def _now_iso() -> str:
    return datetime.now().isoformat()


def _import_status_snapshot() -> dict[str, Any]:
    payload = dict(_MASS_IMPORT_STATUS)
    payload["queued"] = int(payload.get("queued") or 0)
    payload["processed"] = int(payload.get("processed") or 0)
    payload["created"] = int(payload.get("created") or 0)
    payload["failed"] = int(payload.get("failed") or 0)
    payload["chunk_size"] = int(payload.get("chunk_size") or _IMPORT_CHUNK_SIZE)
    payload["running"] = bool(payload.get("running"))
    return payload


def _active_process_tasks() -> int:
    return sum(1 for task in _MASS_TASKS.values() if not task.done())


async def _run_import_queue_worker() -> None:
    global _MASS_IMPORT_TASK
    try:
        while True:
            async with _MASS_IMPORT_LOCK:
                if not _MASS_IMPORT_QUEUE:
                    _MASS_IMPORT_STATUS["running"] = False
                    _MASS_IMPORT_STATUS["queued"] = 0
                    _MASS_IMPORT_STATUS["updated_at"] = _now_iso()
                    if _MASS_IMPORT_STATUS.get("message") == "Queued":
                        _MASS_IMPORT_STATUS["message"] = "Idle"
                    break

                batch = _MASS_IMPORT_QUEUE[:_IMPORT_CHUNK_SIZE]
                del _MASS_IMPORT_QUEUE[:_IMPORT_CHUNK_SIZE]
                _MASS_IMPORT_STATUS["running"] = True
                _MASS_IMPORT_STATUS["queued"] = len(_MASS_IMPORT_QUEUE)
                _MASS_IMPORT_STATUS["updated_at"] = _now_iso()
                _MASS_IMPORT_STATUS["message"] = f"Processing batch ({len(batch)})..."

            created = 0
            failed = 0
            for row in batch:
                # Keep queue pressure bounded so huge imports do not freeze the app.
                while _active_process_tasks() > _MASS_MAX_PENDING_PROCESS_TASKS:
                    await asyncio.sleep(0.08)
                try:
                    state = await _upsert_mass_entry(
                        title=row.get("title") or "",
                        content_type=row.get("type") or "",
                        year=row.get("year") or "",
                        source_note=row.get("source_note") or "csv_excel",
                    )
                    created += 1
                    _schedule_process(str(state.id), mode="full")
                except Exception:
                    failed += 1

            async with _MASS_IMPORT_LOCK:
                _MASS_IMPORT_STATUS["processed"] = int(_MASS_IMPORT_STATUS.get("processed") or 0) + len(batch)
                _MASS_IMPORT_STATUS["created"] = int(_MASS_IMPORT_STATUS.get("created") or 0) + created
                _MASS_IMPORT_STATUS["failed"] = int(_MASS_IMPORT_STATUS.get("failed") or 0) + failed
                _MASS_IMPORT_STATUS["queued"] = len(_MASS_IMPORT_QUEUE)
                _MASS_IMPORT_STATUS["updated_at"] = _now_iso()
                _MASS_IMPORT_STATUS["message"] = "Queued" if _MASS_IMPORT_QUEUE else "Idle"

            await _mass_broadcast_snapshot(force=True)
            await asyncio.sleep(0)

        await _mass_broadcast_snapshot()
    finally:
        async with _MASS_IMPORT_LOCK:
            _MASS_IMPORT_TASK = None
            if not _MASS_IMPORT_QUEUE:
                _MASS_IMPORT_STATUS["running"] = False
                _MASS_IMPORT_STATUS["queued"] = 0
                _MASS_IMPORT_STATUS["updated_at"] = _now_iso()
                if _MASS_IMPORT_STATUS.get("message") == "Queued":
                    _MASS_IMPORT_STATUS["message"] = "Idle"


async def _enqueue_import_rows(rows: list[dict]) -> int:
    global _MASS_IMPORT_TASK
    payload = []
    for row in rows:
        payload.append(
            {
                "title": (row.get("title") or "").strip(),
                "type": (row.get("type") or "").strip(),
                "year": (row.get("year") or "").strip(),
                "source_note": "csv_excel",
            }
        )
    payload = [row for row in payload if row.get("title")]
    if not payload:
        return 0

    async with _MASS_IMPORT_LOCK:
        was_idle = (not _MASS_IMPORT_STATUS.get("running")) and (not _MASS_IMPORT_QUEUE)
        if was_idle:
            _MASS_IMPORT_STATUS["processed"] = 0
            _MASS_IMPORT_STATUS["created"] = 0
            _MASS_IMPORT_STATUS["failed"] = 0
            _MASS_IMPORT_STATUS["started_at"] = _now_iso()
        _MASS_IMPORT_QUEUE.extend(payload)
        _MASS_IMPORT_STATUS["running"] = True
        _MASS_IMPORT_STATUS["queued"] = len(_MASS_IMPORT_QUEUE)
        _MASS_IMPORT_STATUS["chunk_size"] = _IMPORT_CHUNK_SIZE
        _MASS_IMPORT_STATUS["updated_at"] = _now_iso()
        _MASS_IMPORT_STATUS["message"] = "Queued"
        if _MASS_IMPORT_TASK is None or _MASS_IMPORT_TASK.done():
            _MASS_IMPORT_TASK = asyncio.create_task(_run_import_queue_worker())

    return len(payload)


def _parse_csv_payload(raw: bytes) -> list[dict]:
    text = ""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except Exception:
            text = ""
    if not text.strip():
        return []

    stream = io.StringIO(text)
    reader = csv.reader(stream)
    all_rows = [list(row) for row in reader if any((cell or "").strip() for cell in row)]
    if not all_rows:
        return []

    header = [str(x or "").strip().lower() for x in all_rows[0]]
    title_idx = -1
    type_idx = -1
    year_idx = -1
    for idx, col in enumerate(header):
        if col in {"title", "name", "content", "content_name"} and title_idx < 0:
            title_idx = idx
        if col in {"type", "content_type", "category"} and type_idx < 0:
            type_idx = idx
        if col in {"year", "release_year"} and year_idx < 0:
            year_idx = idx

    start_row = 1 if title_idx >= 0 else 0
    if title_idx < 0:
        title_idx = 0

    out: list[dict] = []
    for row in all_rows[start_row:]:
        if title_idx >= len(row):
            continue
        title = str(row[title_idx] or "").strip()
        if not title:
            continue
        type_hint = str(row[type_idx] or "").strip().lower() if type_idx >= 0 and type_idx < len(row) else ""
        year = str(row[year_idx] or "").strip() if year_idx >= 0 and year_idx < len(row) else ""
        out.append({"title": title, "type": type_hint, "year": year})
    return out


def _parse_xlsx_payload(raw: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise ValueError("XLSX parsing requires openpyxl. Add openpyxl to requirements.") from exc

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    matrix: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        values = [str(x or "").strip() for x in row]
        if any(values):
            matrix.append(values)
    if not matrix:
        return []

    header = [str(x).lower() for x in matrix[0]]
    title_idx = -1
    type_idx = -1
    year_idx = -1
    for idx, col in enumerate(header):
        if col in {"title", "name", "content", "content_name"} and title_idx < 0:
            title_idx = idx
        if col in {"type", "content_type", "category"} and type_idx < 0:
            type_idx = idx
        if col in {"year", "release_year"} and year_idx < 0:
            year_idx = idx

    start_row = 1 if title_idx >= 0 else 0
    if title_idx < 0:
        title_idx = 0

    out: list[dict] = []
    for row in matrix[start_row:]:
        if title_idx >= len(row):
            continue
        title = str(row[title_idx] or "").strip()
        if not title:
            continue
        type_hint = str(row[type_idx] or "").strip().lower() if type_idx >= 0 and type_idx < len(row) else ""
        year = str(row[year_idx] or "").strip() if year_idx >= 0 and year_idx < len(row) else ""
        out.append({"title": title, "type": type_hint, "year": year})
    return out


def _dedupe_rows(rows: list[dict]) -> list[dict]:
    out: list[dict] = []
    seen: set[str] = set()
    for row in rows:
        title = (row.get("title") or "").strip()
        if not title:
            continue
        ctype = _guess_type(title, row.get("type") or "")
        key = f"{ctype}:{_title_signature(title)}"
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "title": title,
            "type": ctype,
            "year": (row.get("year") or "").strip(),
        })
    return out


async def _fetch_storage_candidates(title: str) -> list[FileSystemItem]:
    return await _fetch_storage_candidates_multi([title])


def _search_titles_for_state(state: MassContentState) -> list[str]:
    values: list[str] = []
    values.extend(_extract_title_markers(list(state.source_inputs or [])))
    if state.title:
        values.append(state.title)
    expanded: list[str] = []
    for v in values:
        expanded.extend(_expand_title_variants(v))
    # Unique preserve order
    seen: set[str] = set()
    out: list[str] = []
    for v in expanded:
        key = _norm_title(v)
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(v)
    return out


async def _get_storage_pool_cached() -> list[FileSystemItem]:
    now = datetime.now()
    expires_at = _STORAGE_POOL_CACHE.get("expires_at") or datetime.min
    rows = _STORAGE_POOL_CACHE.get("rows") or []
    if rows and now < expires_at:
        return rows
    query = {
        "is_folder": False,
        "$or": [
            {"source": "bot"},
            {"source": "storage", "catalog_status": {"$nin": ["published", "used"]}},
            {"source": "upload", "catalog_status": {"$nin": ["published", "used"]}},
        ],
    }
    fresh = await FileSystemItem.find(query).sort("-created_at").limit(5500).to_list()
    _STORAGE_POOL_CACHE["rows"] = fresh
    _STORAGE_POOL_CACHE["expires_at"] = datetime.fromtimestamp(now.timestamp() + _STORAGE_POOL_TTL_SECONDS)
    return fresh


async def _fetch_storage_candidates_multi(search_titles: list[str]) -> list[FileSystemItem]:
    expanded: list[str] = []
    for title in search_titles:
        expanded.extend(_expand_title_variants(title))
    # Deduplicate and keep meaningful titles
    seen: set[str] = set()
    title_variants: list[str] = []
    for title in expanded:
        key = _norm_title(title)
        if not key or key in seen:
            continue
        seen.add(key)
        title_variants.append(title)

    if not title_variants:
        return []

    search_tokens = set()
    for title in title_variants:
        search_tokens.update(_clean_match_tokens(title, for_file_name=False))
    # Prefer meaningful tokens only for broad fallback prefilter.
    strong_tokens = sorted([t for t in search_tokens if len(t) >= 3], key=len, reverse=True)

    # First pass: database regex match for efficiency.
    name_or: list[dict] = []
    for title in title_variants[:10]:
        pattern = _build_title_regex(title)
        if not pattern:
            continue
        name_or.append({"name": {"$regex": pattern, "$options": "i"}})
        name_or.append({"title": {"$regex": pattern, "$options": "i"}})
        name_or.append({"series_title": {"$regex": pattern, "$options": "i"}})

    rows: list[FileSystemItem] = []
    if name_or:
        query = {
            "is_folder": False,
            "$and": [
                {"$or": name_or},
                {
                    "$or": [
                        {"source": "bot"},
                        {"source": "storage", "catalog_status": {"$nin": ["published", "used"]}},
                        {"source": "upload", "catalog_status": {"$nin": ["published", "used"]}},
                    ]
                },
            ],
        }
        rows = await FileSystemItem.find(query).sort("-created_at").limit(3500).to_list()

    # Fallback pass: broader scan if strict query produced very few hits.
    if len(rows) < 3:
        fallback_rows = await _get_storage_pool_cached()
        if fallback_rows:
            # Merge by id
            merged = {str(x.id): x for x in rows}
            for item in fallback_rows:
                if strong_tokens:
                    name_l = (item.name or "").lower()
                    if not any(tok in name_l for tok in strong_tokens[:4]):
                        continue
                merged[str(item.id)] = item
            rows = list(merged.values())

    filtered = [
        row for row in rows
        if _is_video_row(row)
        and _title_match(
            title_variants,
            row.name or "",
            row_title=(getattr(row, "title", "") or ""),
            row_series_title=(getattr(row, "series_title", "") or ""),
        )
    ]
    # Keep newest first
    filtered.sort(key=lambda x: getattr(x, "created_at", None) or datetime.min, reverse=True)
    return filtered


def _best_row(
    rows: list[dict],
    season: int,
    episode: int,
    quality: str | None = None,
    preferred_file_id: str | None = None,
) -> dict | None:
    picks = [r for r in rows if int(r.get("season") or 0) == season and int(r.get("episode") or 0) == episode]
    if quality:
        q = (quality or "").upper()
        picks = [r for r in picks if (r.get("quality") or "").upper() == q]
    if not picks:
        return None
    if preferred_file_id:
        pref = (preferred_file_id or "").strip()
        for row in picks:
            if str(row.get("file_id") or "").strip() == pref:
                return row
    picks.sort(
        key=lambda x: (
            _quality_rank(x.get("quality") or ""),
            1 if (x.get("source") or "").lower() == "storage" else 0,
            int(x.get("size") or 0),
        ),
        reverse=True,
    )
    return picks[0]


def _series_choice_bucket(season: int, episode: int, quality: str) -> str:
    return f"series:s{int(season)}:e{int(episode)}:q{(quality or 'HD').upper()}"


def _movie_choice_bucket(quality: str) -> str:
    return f"movie:q{(quality or 'HD').upper()}"


async def _scan_files_for_state(
    state: MassContentState,
    prebuilt_found_rows: list[dict] | None = None,
) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict], str, str, bool]:
    content_year = _content_release_year(state)

    quality_overrides: dict[str, str] = {}
    for file_id, value in dict(getattr(state, "file_quality_overrides", {}) or {}).items():
        key = str(file_id or "").strip()
        if not key:
            continue
        quality_overrides[key] = _normalize_quality_label(str(value or ""))

    series_overrides: dict[str, dict[str, int]] = {}
    raw_series_overrides = dict(getattr(state, "file_season_episode_overrides", {}) or {})
    for file_id, payload in raw_series_overrides.items():
        key = str(file_id or "").strip()
        if not key or not isinstance(payload, dict):
            continue
        season_val = _int_or_none(payload.get("season"))
        episode_val = _int_or_none(payload.get("episode"))
        if season_val and episode_val:
            series_overrides[key] = {"season": season_val, "episode": episode_val}

    found_rows: list[dict] = []
    if prebuilt_found_rows is None:
        search_titles = _search_titles_for_state(state)
        candidates = await _fetch_storage_candidates_multi(search_titles or [state.title])

        # Manual include pool: admin can attach specific storage files even if title matching misses them.
        include_ids = {
            str(x).strip()
            for x in (getattr(state, "included_file_ids", []) or [])
            if str(x).strip()
        }
        if include_ids:
            include_rows = await FileSystemItem.find(
                In(FileSystemItem.id, _cast_ids(list(include_ids)))
            ).to_list()
            merged_candidates: dict[str, FileSystemItem] = {
                str(row.id): row for row in candidates
            }
            for row in include_rows:
                if not _is_video_row(row):
                    continue
                merged_candidates[str(row.id)] = row
            candidates = list(merged_candidates.values())

        for row in candidates:
            file_name = row.name or ""
            # Year rule:
            # - If filename has no year token -> allow.
            # - If filename has year token(s) and content year known -> must include content year.
            file_years = _extract_years_from_text(file_name)
            if content_year and file_years and content_year not in file_years:
                continue

            file_id = str(row.id)
            if state.content_type == "series":
                season, episode = _series_season_episode(file_name)
                override_se = series_overrides.get(file_id) or {}
                season = int(override_se.get("season") or season or 0)
                episode = int(override_se.get("episode") or episode or 0)
                if season <= 0 or episode <= 0:
                    continue
                quality = quality_overrides.get(file_id) or _series_quality_from_name(file_name)
                found_rows.append(
                    {
                        "file_id": file_id,
                        "name": file_name,
                        "size": int(row.size or 0),
                        "quality": quality,
                        "season": int(season),
                        "episode": int(episode),
                        "source": row.source or "",
                        "source_label": _source_label(row),
                        "upload_label": _source_upload_label(row),
                    }
                )
            else:
                quality = quality_overrides.get(file_id) or _movie_quality_from_name_or_size(
                    file_name,
                    int(row.size or 0),
                )
                found_rows.append(
                    {
                        "file_id": file_id,
                        "name": file_name,
                        "size": int(row.size or 0),
                        "quality": quality,
                        "season": None,
                        "episode": None,
                        "source": row.source or "",
                        "source_label": _source_label(row),
                        "upload_label": _source_upload_label(row),
                    }
                )
    else:
        for row in list(prebuilt_found_rows or []):
            if not isinstance(row, dict):
                continue
            file_id = str(row.get("file_id") or "").strip()
            if not file_id:
                continue
            file_name = str(row.get("name") or "").strip()
            file_size = int(row.get("size") or 0)
            src = str(row.get("source") or "").strip().lower()
            source_label = str(row.get("source_label") or "").strip()
            upload_label = str(row.get("upload_label") or "").strip()
            if not source_label:
                source_label = "Telegram Storage" if src == "storage" else ("Bot Listening" if src == "bot" else "Storage")
            if not upload_label:
                upload_label = "Uploaded from Telegram Storage" if src == "storage" else ("Found by Bot Listener" if src == "bot" else "Matched")

            # Year rule:
            # - If filename has no year token -> allow.
            # - If filename has year token(s) and content year known -> must include content year.
            file_years = _extract_years_from_text(file_name)
            if content_year and file_years and content_year not in file_years:
                continue

            if state.content_type == "series":
                base_season = _int_or_none(row.get("season"))
                base_episode = _int_or_none(row.get("episode"))
                if not base_season or not base_episode:
                    guessed_season, guessed_episode = _series_season_episode(file_name)
                    base_season = base_season or guessed_season
                    base_episode = base_episode or guessed_episode
                override_se = series_overrides.get(file_id) or {}
                season = int(override_se.get("season") or base_season or 0)
                episode = int(override_se.get("episode") or base_episode or 0)
                if season <= 0 or episode <= 0:
                    continue
                quality = quality_overrides.get(file_id) or _normalize_quality_label(row.get("quality") or "") or _series_quality_from_name(file_name)
                found_rows.append(
                    {
                        "file_id": file_id,
                        "name": file_name,
                        "size": file_size,
                        "quality": quality,
                        "season": int(season),
                        "episode": int(episode),
                        "source": src,
                        "source_label": source_label,
                        "upload_label": upload_label,
                    }
                )
            else:
                quality = quality_overrides.get(file_id) or _normalize_quality_label(row.get("quality") or "") or _movie_quality_from_name_or_size(
                    file_name,
                    file_size,
                )
                found_rows.append(
                    {
                        "file_id": file_id,
                        "name": file_name,
                        "size": file_size,
                        "quality": quality,
                        "season": None,
                        "episode": None,
                        "source": src,
                        "source_label": source_label,
                        "upload_label": upload_label,
                    }
                )

    # Deduplicate rows by file id.
    uniq: dict[str, dict] = {}
    for row in found_rows:
        uniq[row["file_id"]] = row
    found_rows = list(uniq.values())

    excluded_ids = {str(x).strip() for x in (getattr(state, "excluded_file_ids", []) or []) if str(x).strip()}
    if excluded_ids:
        found_rows = [row for row in found_rows if str(row.get("file_id") or "").strip() not in excluded_ids]

    selected_file_map = dict(getattr(state, "selected_file_map", {}) or {})
    file_choice_groups: list[dict] = []

    notes: list[dict] = []
    missing: list[dict] = []
    upload_plan: list[dict] = []
    panel = "file_not_found"
    file_status = "missing"
    upload_ready = False

    if state.content_type == "movie":
        by_quality_rows: dict[str, list[dict]] = {}
        for row in found_rows:
            quality = (row.get("quality") or "HD").upper()
            by_quality_rows.setdefault(quality, []).append(row)

        by_quality: dict[str, dict] = {}
        for quality, rows in by_quality_rows.items():
            bucket = _movie_choice_bucket(quality)
            preferred_file_id = str(selected_file_map.get(bucket) or "").strip()
            chosen = None
            if preferred_file_id:
                for cand in rows:
                    if str(cand.get("file_id") or "").strip() == preferred_file_id:
                        chosen = cand
                        break
            if not chosen:
                rows_sorted = sorted(
                    rows,
                    key=lambda x: (
                        1 if (x.get("source") or "").lower() == "storage" else 0,
                        int(x.get("size") or 0),
                    ),
                    reverse=True,
                )
                chosen = rows_sorted[0] if rows_sorted else None
            if chosen:
                by_quality[quality] = chosen

            if len(rows) > 1:
                candidates_sorted = sorted(
                    rows,
                    key=lambda x: (
                        1 if str(x.get("file_id") or "").strip() == preferred_file_id else 0,
                        1 if (x.get("source") or "").lower() == "storage" else 0,
                        int(x.get("size") or 0),
                    ),
                    reverse=True,
                )
                selected_id = preferred_file_id or (str(candidates_sorted[0].get("file_id") or "").strip() if candidates_sorted else "")
                file_choice_groups.append(
                    {
                        "bucket_key": bucket,
                        "label": f"{quality}",
                        "season": None,
                        "episode": None,
                        "quality": quality,
                        "selected_file_id": selected_id,
                        "candidates": [
                            {
                                "file_id": str(c.get("file_id") or ""),
                                "name": c.get("name") or "",
                                "size": int(c.get("size") or 0),
                                "size_label": format_size(int(c.get("size") or 0)),
                                "source_label": c.get("source_label") or "",
                                "selected": str(c.get("file_id") or "").strip() == selected_id,
                            }
                            for c in candidates_sorted
                        ],
                    }
                )

        required = ["1080P", "720P", "480P"]
        for quality in required:
            picked = by_quality.get(quality)
            if picked:
                notes.append(
                    {
                        "season": None,
                        "episode": None,
                        "quality": quality,
                        "state": "found",
                        "text": f"{quality} -> Found ({picked.get('upload_label')}){_file_note_suffix(picked)}",
                    }
                )
            else:
                notes.append(
                    {
                        "season": None,
                        "episode": None,
                        "quality": quality,
                        "state": "missing",
                        "text": f"{quality} -> NOT FOUND (Upload Required)",
                    }
                )
                missing.append({"quality": quality, "note": f"Missing movie quality {quality}"})

        sorted_found = sorted(by_quality.values(), key=lambda x: _quality_rank(x.get("quality") or ""), reverse=True)[:4]
        for row in sorted_found:
            upload_plan.append(
                {
                    "file_id": row.get("file_id"),
                    "quality": (row.get("quality") or "HD").upper(),
                    "season": None,
                    "episode": None,
                    "episode_title": "",
                }
            )

        if not found_rows:
            panel = "file_not_found"
            file_status = "missing"
            upload_ready = False
        elif not missing:
            panel = "complete"
            file_status = "complete"
            upload_ready = True
        else:
            panel = "incomplete"
            file_status = "incomplete"
            upload_ready = False
        return found_rows, notes, missing, upload_plan, file_choice_groups, panel, file_status, upload_ready

    seasons = list(state.seasons or [])
    if not seasons:
        # If TMDB has no seasons payload, infer from file names.
        grouped: dict[int, set[int]] = {}
        for row in found_rows:
            season = int(row.get("season") or 0)
            episode = int(row.get("episode") or 0)
            if season <= 0 or episode <= 0:
                continue
            grouped.setdefault(season, set()).add(episode)
        seasons = [
            {
                "season": season,
                "name": f"Season {season}",
                "episodes": [{"episode": ep, "name": ""} for ep in sorted(list(episodes))],
            }
            for season, episodes in sorted(grouped.items())
        ]

    season_ready: list[bool] = []
    for season_row in seasons:
        season_no = int(season_row.get("season") or 0)
        episode_rows = list(season_row.get("episodes") or [])
        expected_episodes = sorted([int(x.get("episode") or 0) for x in episode_rows if int(x.get("episode") or 0) > 0])
        if not expected_episodes:
            continue

        season_found = [r for r in found_rows if int(r.get("season") or 0) == season_no]
        quality_to_episodes: dict[str, set[int]] = {}
        for row in season_found:
            quality = (row.get("quality") or "HD").upper()
            quality_to_episodes.setdefault(quality, set()).add(int(row.get("episode") or 0))

        # Build duplicate-choice groups for admin (same season/episode/quality with multiple files).
        season_bucket_rows: dict[str, list[dict]] = {}
        for row in season_found:
            episode_no = int(row.get("episode") or 0)
            quality = (row.get("quality") or "HD").upper()
            if episode_no <= 0:
                continue
            bucket_key = _series_choice_bucket(season_no, episode_no, quality)
            season_bucket_rows.setdefault(bucket_key, []).append(row)

        for bucket_key, rows in season_bucket_rows.items():
            if len(rows) <= 1:
                continue
            preferred_file_id = str(selected_file_map.get(bucket_key) or "").strip()
            candidates_sorted = sorted(
                rows,
                key=lambda x: (
                    1 if str(x.get("file_id") or "").strip() == preferred_file_id else 0,
                    1 if (x.get("source") or "").lower() == "storage" else 0,
                    int(x.get("size") or 0),
                ),
                reverse=True,
            )
            first = candidates_sorted[0] if candidates_sorted else {}
            selected_id = preferred_file_id or str(first.get("file_id") or "").strip()
            parts = bucket_key.split(":")
            episode_from_bucket = int(parts[2].replace("e", "")) if len(parts) >= 3 else 0
            quality_from_bucket = parts[3].replace("q", "").upper() if len(parts) >= 4 else ""
            file_choice_groups.append(
                {
                    "bucket_key": bucket_key,
                    "label": f"S{season_no:02d}E{episode_from_bucket:02d} {quality_from_bucket}",
                    "season": season_no,
                    "episode": episode_from_bucket,
                    "quality": quality_from_bucket,
                    "selected_file_id": selected_id,
                    "candidates": [
                        {
                            "file_id": str(c.get("file_id") or ""),
                            "name": c.get("name") or "",
                            "size": int(c.get("size") or 0),
                            "size_label": format_size(int(c.get("size") or 0)),
                            "source_label": c.get("source_label") or "",
                            "selected": str(c.get("file_id") or "").strip() == selected_id,
                        }
                        for c in candidates_sorted
                    ],
                }
            )

        complete_qualities = []
        for quality, covered in quality_to_episodes.items():
            if all(ep in covered for ep in expected_episodes):
                complete_qualities.append(quality)

        chosen_quality = "1080P"
        if complete_qualities:
            complete_qualities.sort(key=_quality_rank, reverse=True)
            chosen_quality = complete_qualities[0]
        elif quality_to_episodes:
            ranked = sorted(
                quality_to_episodes.items(),
                key=lambda kv: (len(kv[1]), _quality_rank(kv[0])),
                reverse=True,
            )
            chosen_quality = ranked[0][0]

        if len(complete_qualities) > 4:
            complete_qualities = sorted(complete_qualities, key=_quality_rank, reverse=True)[:4]
        season_ready.append(len(complete_qualities) > 0)
        # Keep multiple qualities per season (max 4):
        # 1) all fully-complete qualities, then
        # 2) strongest partial qualities for extra episode-level captures.
        ranked_qualities = sorted(
            quality_to_episodes.items(),
            key=lambda kv: (len(kv[1]), _quality_rank(kv[0])),
            reverse=True,
        )
        season_quality_targets = list(complete_qualities)
        for quality, _covered in ranked_qualities:
            if quality in season_quality_targets:
                continue
            if len(season_quality_targets) >= 4:
                break
            season_quality_targets.append(quality)
        if not season_quality_targets and chosen_quality:
            season_quality_targets = [chosen_quality]
        complete_quality_set = set(complete_qualities)

        for episode_no in expected_episodes:
            episode_title = ""
            for entry in episode_rows:
                if int(entry.get("episode") or 0) == episode_no:
                    episode_title = (entry.get("name") or "").strip()
                    break

            if complete_qualities:
                # Add one plan row per complete quality for the same episode.
                for quality in complete_qualities:
                    bucket_key = _series_choice_bucket(season_no, episode_no, quality)
                    preferred_file_id = str(selected_file_map.get(bucket_key) or "").strip()
                    same_quality = _best_row(season_found, season_no, episode_no, quality, preferred_file_id=preferred_file_id)
                    if same_quality:
                        notes.append(
                            {
                                "season": season_no,
                                "episode": episode_no,
                                "quality": quality,
                                "state": "found",
                                "text": f"Season {season_no} Episode {episode_no} {quality} -> Found ({same_quality.get('upload_label')}){_file_note_suffix(same_quality)}",
                            }
                        )
                        upload_plan.append(
                            {
                                "file_id": same_quality.get("file_id"),
                                "quality": quality,
                                "season": season_no,
                                "episode": episode_no,
                                "episode_title": episode_title,
                            }
                        )
                    else:
                        # Defensive note; should be rare because quality is marked complete.
                        notes.append(
                            {
                                "season": season_no,
                                "episode": episode_no,
                                "quality": quality,
                                "state": "missing",
                                "text": f"Season {season_no} Episode {episode_no} {quality} -> NOT FOUND (Upload Required)",
                            }
                        )
                        missing.append(
                            {
                                "season": season_no,
                                "episode": episode_no,
                                "quality": quality,
                                "note": f"Season {season_no} Episode {episode_no} missing at {quality}",
                            }
                        )
                # Also capture extra partial qualities when available for this episode
                # (do not mark missing if absent, because they are optional extras).
                for quality in season_quality_targets:
                    if quality in complete_quality_set:
                        continue
                    bucket_key = _series_choice_bucket(season_no, episode_no, quality)
                    preferred_file_id = str(selected_file_map.get(bucket_key) or "").strip()
                    extra_quality_row = _best_row(season_found, season_no, episode_no, quality, preferred_file_id=preferred_file_id)
                    if not extra_quality_row:
                        continue
                    notes.append(
                        {
                            "season": season_no,
                            "episode": episode_no,
                            "quality": quality,
                            "state": "found_partial",
                            "text": f"Season {season_no} Episode {episode_no} {quality} -> Found ({extra_quality_row.get('upload_label')}){_file_note_suffix(extra_quality_row)}",
                        }
                    )
                    upload_plan.append(
                        {
                            "file_id": extra_quality_row.get("file_id"),
                            "quality": quality,
                            "season": season_no,
                            "episode": episode_no,
                            "episode_title": episode_title,
                        }
                    )
                continue

            chosen_bucket = _series_choice_bucket(season_no, episode_no, chosen_quality)
            preferred_file_id = str(selected_file_map.get(chosen_bucket) or "").strip()
            same_quality = _best_row(season_found, season_no, episode_no, chosen_quality, preferred_file_id=preferred_file_id)
            any_quality = _best_row(season_found, season_no, episode_no, None)
            if same_quality:
                notes.append(
                    {
                        "season": season_no,
                        "episode": episode_no,
                        "quality": chosen_quality,
                        "state": "found",
                        "text": f"Season {season_no} Episode {episode_no} {chosen_quality} -> Found ({same_quality.get('upload_label')}){_file_note_suffix(same_quality)}",
                    }
                )
                upload_plan.append(
                    {
                        "file_id": same_quality.get("file_id"),
                        "quality": chosen_quality,
                        "season": season_no,
                        "episode": episode_no,
                        "episode_title": episode_title,
                    }
                )
            elif any_quality:
                quality = (any_quality.get("quality") or "HD").upper()
                notes.append(
                    {
                        "season": season_no,
                        "episode": episode_no,
                        "quality": quality,
                        "state": "found_partial",
                        "text": f"Season {season_no} Episode {episode_no} {quality} -> Found ({any_quality.get('upload_label')}){_file_note_suffix(any_quality)}",
                    }
                )
                upload_plan.append(
                    {
                        "file_id": any_quality.get("file_id"),
                        "quality": quality,
                        "season": season_no,
                        "episode": episode_no,
                        "episode_title": episode_title,
                    }
                )
                missing.append(
                    {
                        "season": season_no,
                        "episode": episode_no,
                        "quality": chosen_quality,
                        "note": f"Season {season_no} Episode {episode_no} missing at {chosen_quality}",
                    }
                )
            else:
                notes.append(
                    {
                        "season": season_no,
                        "episode": episode_no,
                        "quality": chosen_quality,
                        "state": "missing",
                        "text": f"Season {season_no} Episode {episode_no} {chosen_quality} -> NOT FOUND (Upload Required)",
                    }
                )
                missing.append(
                    {
                        "season": season_no,
                        "episode": episode_no,
                        "quality": chosen_quality,
                        "note": f"Season {season_no} Episode {episode_no} missing",
                    }
                )

    # Deduplicate upload plan by file id while preserving order.
    seen_ids: set[str] = set()
    clean_upload_plan: list[dict] = []
    for row in upload_plan:
        file_id = (row.get("file_id") or "").strip()
        if not file_id or file_id in seen_ids:
            continue
        seen_ids.add(file_id)
        clean_upload_plan.append(row)
    upload_plan = clean_upload_plan

    if not found_rows:
        panel = "file_not_found"
        file_status = "missing"
        upload_ready = False
    elif season_ready and all(season_ready):
        panel = "complete"
        file_status = "complete"
        upload_ready = True
    else:
        panel = "incomplete"
        file_status = "incomplete"
        upload_ready = False

    return found_rows, notes, missing, upload_plan, file_choice_groups, panel, file_status, upload_ready


async def _recompute_mass_item_files_fast(
    row: MassContentState,
    base_found_rows: list[dict] | None = None,
) -> None:
    found_rows, notes, missing, upload_plan, file_choice_groups, panel, file_status, upload_ready = await _scan_files_for_state(
        row,
        prebuilt_found_rows=list(base_found_rows or []),
    )
    row.matched_files = found_rows
    row.live_notes = notes
    row.missing_items = missing
    row.upload_plan = upload_plan
    row.file_choice_groups = file_choice_groups
    row.panel = panel
    row.file_status = file_status
    row.upload_ready = bool(upload_ready)
    row.uploaded = False
    row.uploaded_at = None
    row.upload_state = "idle"
    row.upload_message = ""
    row.last_error = None
    row.updated_at = datetime.now()
    await row.save()
    await _mass_broadcast_snapshot()


async def _upsert_mass_entry(title: str, content_type: str, year: str, source_note: str) -> MassContentState:
    clean_title = " ".join((title or "").split()).strip()
    if not clean_title:
        raise ValueError("Missing title")
    guessed_type = _guess_type(clean_title, content_type)
    key = f"{guessed_type}:{_norm_title(clean_title)}"
    title_sig = _title_signature(clean_title)

    existing = await MassContentState.find_one(MassContentState.key == key)
    # Fallback: match existing records by canonical token signature so reordered/symbol-heavy
    # duplicates map into one row (example: "A+B: C" vs "A B C").
    if not existing and title_sig:
        recent_same_type = await MassContentState.find(MassContentState.content_type == guessed_type).sort("-updated_at").limit(900).to_list()
        for row in recent_same_type:
            if _title_signature(row.title or "") == title_sig:
                existing = row
                break
    if existing:
        changed = False
        if year and not existing.year:
            existing.year = year
            changed = True
        markers: list[str] = []
        source_marker = _to_source_marker(source_note)
        title_marker = _to_title_marker(clean_title)
        if source_marker:
            markers.append(source_marker)
        if title_marker:
            markers.append(title_marker)
        merged_inputs = list(existing.source_inputs or [])
        for marker in markers:
            if marker and marker not in merged_inputs:
                merged_inputs.append(marker)
                changed = True
        if merged_inputs != list(existing.source_inputs or []):
            existing.source_inputs = merged_inputs
            changed = True
        if existing.panel in {"tmdb_not_found", "file_not_found", "incomplete", "skipped"}:
            existing.panel = "processing"
            existing.tmdb_status = "pending"
            existing.file_status = "pending"
            existing.last_error = None
            changed = True
        if changed:
            existing.updated_at = datetime.now()
            await existing.save()
        return existing

    row = MassContentState(
        key=key,
        title=clean_title,
        normalized_title=_norm_title(clean_title),
        content_type=guessed_type,
        year=(year or "").strip() or None,
        panel="processing",
        tmdb_status="pending",
        file_status="pending",
        source_inputs=[x for x in [_to_source_marker(source_note), _to_title_marker(clean_title)] if x],
        created_at=datetime.now(),
        updated_at=datetime.now(),
    )
    await row.insert()
    return row


async def _content_exists_for_input(title: str, content_type: str, year: str = "") -> tuple[bool, dict]:
    clean_title = " ".join((title or "").split()).strip()
    if not clean_title:
        return False, {}
    ctype = _guess_type(clean_title, content_type)
    norm_title = _norm_title(clean_title)
    if not norm_title:
        return False, {}

    query: dict[str, Any] = {
        "content_type": ctype,
        "status": "published",
        "title": {"$regex": f"^{re.escape(clean_title)}$", "$options": "i"},
    }
    doc = await ContentItem.find_one(query)
    if not doc:
        # Fast path on indexed normalized title (search_title).
        doc = await ContentItem.find_one({
            "content_type": ctype,
            "status": "published",
            "search_title": norm_title,
        })
    if not doc:
        # Fallback: fuzzy regex over indexed search_title.
        token_regex = _build_title_regex(clean_title)
        if token_regex:
            doc = await ContentItem.find_one({
                "content_type": ctype,
                "status": "published",
                "search_title": {"$regex": token_regex, "$options": "i"},
            })
    if not doc:
        return False, {}
    return True, {
        "id": str(doc.id),
        "title": (doc.title or "").strip(),
        "year": (doc.year or "").strip(),
        "content_type": (doc.content_type or ctype).strip().lower(),
        "slug": (doc.slug or "").strip(),
    }


async def _build_existing_content_index() -> dict[str, set]:
    by_title: set[tuple[str, str]] = set()
    details_by_title: dict[tuple[str, str], dict[str, str]] = {}
    projection = {"title": 1, "content_type": 1, "year": 1, "slug": 1}
    cursor = _collection_for(ContentItem).find({"status": "published"}, projection).limit(50000)
    async for doc in cursor:
        ctype = (str(doc.get("content_type") or "movie")).strip().lower()
        title = (str(doc.get("title") or "")).strip()
        norm = _norm_title(title)
        if not norm:
            continue
        key = (ctype, norm)
        by_title.add(key)
        if key not in details_by_title:
            details_by_title[key] = {
                "id": str(doc.get("_id") or ""),
                "title": title,
                "year": (str(doc.get("year") or "")).strip(),
                "content_type": ctype,
                "slug": (str(doc.get("slug") or "")).strip(),
            }
    return {
        "title_only": by_title,
        "details_by_title": details_by_title,
    }


def _row_exists_in_content_index(title: str, content_type: str, year: str, index: dict[str, set]) -> bool:
    ctype = _guess_type(title, content_type)
    norm = _norm_title(title)
    if not norm:
        return False
    return (ctype, norm) in index.get("title_only", set())


def _existing_detail_from_index(title: str, content_type: str, index: dict[str, set]) -> dict:
    ctype = _guess_type(title, content_type)
    norm = _norm_title(title)
    if not norm:
        return {}
    details = index.get("details_by_title", {})
    if not isinstance(details, dict):
        return {}
    payload = details.get((ctype, norm)) or {}
    return dict(payload) if isinstance(payload, dict) else {}


async def _upsert_skipped_entry(
    title: str,
    content_type: str,
    year: str,
    reason: str,
    existing_content: dict | None = None,
) -> MassContentState:
    clean_title = " ".join((title or "").split()).strip()
    if not clean_title:
        raise ValueError("Missing title")
    guessed_type = _guess_type(clean_title, content_type)
    key = f"skipped:{guessed_type}:{_norm_title(clean_title)}"
    row = await MassContentState.find_one(MassContentState.key == key)
    now = datetime.now()
    payload = dict(existing_content or {})

    if row:
        row.title = clean_title
        row.content_type = guessed_type
        if (year or "").strip():
            row.year = (year or "").strip()
        row.panel = "skipped"
        row.tmdb_status = "pending"
        row.file_status = "pending"
        row.upload_ready = False
        row.uploaded = False
        row.uploaded_at = None
        row.upload_state = "idle"
        row.upload_message = ""
        row.skip_reason = (reason or "").strip() or "Already exists in content database."
        row.existing_content = payload
        row.last_error = None
        row.updated_at = now
        await row.save()
        return row

    row = MassContentState(
        key=key,
        title=clean_title,
        normalized_title=_norm_title(clean_title),
        content_type=guessed_type,
        year=(year or "").strip() or None,
        panel="skipped",
        tmdb_status="pending",
        file_status="pending",
        upload_ready=False,
        uploaded=False,
        uploaded_at=None,
        upload_state="idle",
        upload_message="",
        source_inputs=[_to_source_marker("csv_excel"), _to_title_marker(clean_title)],
        skip_reason=(reason or "").strip() or "Already exists in content database.",
        existing_content=payload,
        created_at=now,
        updated_at=now,
    )
    await row.insert()
    return row


async def _cleanup_mass_state_duplicates() -> int:
    global _MASS_DEDUPE_LAST_RUN_TS
    now_ts = asyncio.get_event_loop().time()
    if (now_ts - float(_MASS_DEDUPE_LAST_RUN_TS or 0.0)) < _MASS_DEDUPE_COOLDOWN_SEC:
        return 0
    _MASS_DEDUPE_LAST_RUN_TS = now_ts

    try:
        rows = await asyncio.wait_for(
            MassContentState.find_all().sort("-updated_at").limit(_MASS_DEDUPE_SCAN_LIMIT).to_list(),
            timeout=_MASS_DEDUPE_MAX_RUNTIME_SEC,
        )
    except Exception as exc:
        logger.warning("Mass dedupe skipped due to DB timeout/error: %s", exc)
        return 0
    if not rows:
        return 0
    picked: dict[str, MassContentState] = {}
    delete_ids: set[str] = set()
    for row in rows:
        identity = _row_identity_key(row)
        existing = picked.get(identity)
        if not existing:
            picked[identity] = row
            continue
        if _is_better_row(row, existing):
            delete_ids.add(str(existing.id))
            picked[identity] = row
        else:
            delete_ids.add(str(row.id))
    if not delete_ids:
        return 0
    await MassContentState.find(In(MassContentState.id, _cast_ids(list(delete_ids)))).delete()
    return len(delete_ids)


def _schedule_mass_dedupe() -> None:
    global _MASS_BG_DEDUPE_TASK
    try:
        if _MASS_BG_DEDUPE_TASK is not None and not _MASS_BG_DEDUPE_TASK.done():
            return
    except Exception:
        pass

    async def _runner():
        try:
            await _cleanup_mass_state_duplicates()
        except Exception:
            pass

    try:
        _MASS_BG_DEDUPE_TASK = asyncio.create_task(_runner())
    except Exception:
        _MASS_BG_DEDUPE_TASK = None


async def _recover_stale_upload_rows() -> int:
    global _MASS_UPLOAD_RECOVERY_LAST_RUN_TS
    now_ts = asyncio.get_event_loop().time()
    if (now_ts - float(_MASS_UPLOAD_RECOVERY_LAST_RUN_TS or 0.0)) < _MASS_UPLOAD_RECOVERY_COOLDOWN_SEC:
        return 0
    _MASS_UPLOAD_RECOVERY_LAST_RUN_TS = now_ts

    cutoff = datetime.now() - timedelta(seconds=_MASS_UPLOAD_STALE_SEC)
    query = {
        "uploaded": {"$ne": True},
        "upload_state": {"$in": ["queued", "uploading"]},
        "updated_at": {"$lt": cutoff},
    }
    rows = await MassContentState.find(query).sort("-updated_at").limit(_MASS_UPLOAD_RECOVERY_LIMIT).to_list()
    if not rows:
        return 0

    changed = 0
    for row in rows:
        row_id = str(getattr(row, "id", "") or "")
        task = _MASS_UPLOAD_TASKS.get(row_id)
        if task and not task.done():
            continue
        row.upload_state = "failed"
        row.upload_message = "Upload worker interrupted. Click Upload again."
        if not (row.last_error or "").strip():
            row.last_error = "Upload worker interrupted."
        normalized_panel = _normalize_mass_panel_key(getattr(row, "panel", ""), fallback="")
        if normalized_panel in {"", "uploading"}:
            if row.tmdb_status != "found":
                row.panel = "processing"
            else:
                row.panel = "complete" if bool(getattr(row, "upload_ready", False)) else "incomplete"
        row.updated_at = datetime.now()
        try:
            await row.save()
            changed += 1
        except Exception:
            continue

    if changed > 0:
        _invalidate_mass_snapshot_cache()
    return changed


def _schedule_mass_upload_recovery() -> None:
    global _MASS_BG_UPLOAD_RECOVERY_TASK
    try:
        if _MASS_BG_UPLOAD_RECOVERY_TASK is not None and not _MASS_BG_UPLOAD_RECOVERY_TASK.done():
            return
    except Exception:
        pass

    async def _runner():
        try:
            repaired = await _recover_stale_upload_rows()
            if repaired > 0:
                await _mass_broadcast_snapshot(force=True)
        except Exception:
            pass

    try:
        _MASS_BG_UPLOAD_RECOVERY_TASK = asyncio.create_task(_runner())
    except Exception:
        _MASS_BG_UPLOAD_RECOVERY_TASK = None


async def _process_mass_item(item_id: str, mode: str = "full") -> None:
    lock = _MASS_LOCKS.setdefault(item_id, asyncio.Lock())
    async with lock:
        row = await MassContentState.get(item_id)
        if not row:
            return

        row.panel = "processing"
        if mode in {"full", "tmdb"}:
            row.tmdb_status = "pending"
            row.last_error = None
        if mode in {"full", "files"}:
            row.file_status = "pending"
        row.updated_at = datetime.now()
        await row.save()
        await _mass_broadcast_snapshot()

        try:
            if mode in {"full", "tmdb"}:
                search_payload = await _tmdb_search_best(row.title, row.year or "", row.content_type == "series")
                results = list(search_payload.get("results") or [])
                pick = search_payload.get("pick")
                resolved_is_series = bool(search_payload.get("is_series"))
                if not results or not pick:
                    row.tmdb_status = "not_found"
                    row.panel = "tmdb_not_found"
                    row.file_status = "pending"
                    row.upload_ready = False
                    row.updated_at = datetime.now()
                    await row.save()
                    await _mass_broadcast_snapshot(force=True)
                    return

                tmdb_id = pick.get("id")
                if not tmdb_id:
                    row.tmdb_status = "not_found"
                    row.panel = "tmdb_not_found"
                    row.updated_at = datetime.now()
                    await row.save()
                    await _mass_broadcast_snapshot(force=True)
                    return

                row.content_type = "series" if resolved_is_series else "movie"
                details = await _tmdb_details(int(tmdb_id), resolved_is_series)
                if not details:
                    row.tmdb_status = "not_found"
                    row.panel = "tmdb_not_found"
                    row.updated_at = datetime.now()
                    await row.save()
                    await _mass_broadcast_snapshot(force=True)
                    return

                pre_tmdb_title = (row.title or "").strip()
                tmdb_title = details.get("name") if resolved_is_series else details.get("title")
                if tmdb_title:
                    marker = _to_title_marker(pre_tmdb_title)
                    if marker and marker not in (row.source_inputs or []):
                        row.source_inputs = list(row.source_inputs or []) + [marker]
                    row.title = tmdb_title.strip()
                    row.normalized_title = _norm_title(row.title)

                release_date = (details.get("release_date") or details.get("first_air_date") or "").strip()
                year = release_date[:4] if release_date else (row.year or "")
                if year:
                    row.year = year

                poster_path = details.get("poster_path")
                backdrop_path = details.get("backdrop_path")
                row.poster_url = f"https://image.tmdb.org/t/p/w780{poster_path}" if poster_path else (row.poster_url or "")
                row.backdrop_url = f"https://image.tmdb.org/t/p/w1280{backdrop_path}" if backdrop_path else (row.backdrop_url or "")
                row.description = (details.get("overview") or "").strip()
                row.release_date = release_date or (row.release_date or "")
                row.tmdb_id = int(tmdb_id)
                row.genres = [g.get("name") for g in (details.get("genres") or []) if g.get("name")]

                credits = details.get("credits") or {}
                cast_rows = credits.get("cast") or []
                row.actors = [c.get("name") for c in cast_rows[:12] if c.get("name")]
                cast_profiles = []
                for c in cast_rows[:12]:
                    if not c.get("name"):
                        continue
                    profile_path = c.get("profile_path")
                    cast_profiles.append(
                        {
                            "id": c.get("id"),
                            "name": c.get("name"),
                            "role": c.get("character") or "",
                            "image": f"https://image.tmdb.org/t/p/w185{profile_path}" if profile_path else "",
                        }
                    )
                row.cast_profiles = cast_profiles

                director = ""
                for c in credits.get("crew") or []:
                    if c.get("job") == "Director":
                        director = c.get("name") or ""
                        break
                if not director and resolved_is_series:
                    for c in credits.get("crew") or []:
                        if c.get("job") in {"Creator", "Executive Producer"}:
                            director = c.get("name") or ""
                            break
                row.director = director

                trailer_url = ""
                trailer_key = ""
                for v in (details.get("videos", {}) or {}).get("results", []) or []:
                    if v.get("site") == "YouTube" and v.get("type") in {"Trailer", "Teaser"}:
                        trailer_key = v.get("key") or ""
                        if trailer_key:
                            trailer_url = f"https://www.youtube.com/watch?v={trailer_key}"
                        break
                row.trailer_key = trailer_key
                row.trailer_url = trailer_url

                seasons: list[dict] = []
                if resolved_is_series:
                    for season in details.get("seasons") or []:
                        season_no = int(season.get("season_number") or 0)
                        if season_no <= 0:
                            continue
                        season_details = await _tmdb_get(f"/tv/{tmdb_id}/season/{season_no}", {})
                        episodes = []
                        for ep in season_details.get("episodes") or []:
                            ep_no = int(ep.get("episode_number") or 0)
                            if ep_no <= 0:
                                continue
                            episodes.append({"episode": ep_no, "name": (ep.get("name") or "").strip()})
                        episode_count = int(season.get("episode_count") or len(episodes))
                        # Skip placeholder/upcoming seasons that have no episodes yet.
                        if episode_count <= 0 and not episodes:
                            continue
                        seasons.append(
                            {
                                "season": season_no,
                                "name": season.get("name") or f"Season {season_no}",
                                "episode_count": episode_count,
                                "episodes": episodes,
                            }
                        )
                row.seasons = sorted(seasons, key=lambda x: int(x.get("season") or 0))
                row.tmdb_status = "found"

            if row.tmdb_status != "found":
                row.panel = "tmdb_not_found"
                row.updated_at = datetime.now()
                await row.save()
                await _mass_broadcast_snapshot()
                return

            found_rows, notes, missing, upload_plan, file_choice_groups, panel, file_status, upload_ready = await _scan_files_for_state(row)
            row.matched_files = found_rows
            row.live_notes = notes
            row.missing_items = missing
            row.upload_plan = upload_plan
            row.file_choice_groups = file_choice_groups
            row.panel = panel
            row.file_status = file_status
            row.upload_ready = bool(upload_ready)
            row.last_error = None
            row.updated_at = datetime.now()
            await row.save()
            await _mass_broadcast_snapshot()
        except Exception as exc:
            row.panel = "processing"
            row.last_error = str(exc)
            row.updated_at = datetime.now()
            await row.save()
            await _mass_broadcast_snapshot()


def _schedule_process(item_id: str, mode: str = "full") -> None:
    existing = _MASS_TASKS.get(item_id)
    if existing and not existing.done():
        return

    async def _runner():
        try:
            async with _MASS_PROCESS_SEMAPHORE:
                await _process_mass_item(item_id, mode=mode)
        finally:
            _MASS_TASKS.pop(item_id, None)

    _MASS_TASKS[item_id] = asyncio.create_task(_runner())


def _serialize_row(row: MassContentState) -> dict:
    matched_preview = []
    matched_full = []
    for item in list(row.matched_files or [])[:12]:
        season = item.get("season")
        episode = item.get("episode")
        se_label = ""
        if isinstance(season, int) and season > 0 and isinstance(episode, int) and episode > 0:
            se_label = f"S{season:02d}E{episode:02d}"
        matched_preview.append(
            {
                "file_id": str(item.get("file_id") or ""),
                "name": (item.get("name") or "").strip(),
                "quality": (item.get("quality") or "").upper(),
                "se": se_label,
                "size": int(item.get("size") or 0),
                "size_label": format_size(int(item.get("size") or 0)),
                "source_label": item.get("source_label") or "",
            }
        )
    for item in list(row.matched_files or [])[:100]:
        season = _int_or_none(item.get("season"))
        episode = _int_or_none(item.get("episode"))
        matched_full.append(
            {
                "file_id": str(item.get("file_id") or ""),
                "name": (item.get("name") or "").strip(),
                "quality": _normalize_quality_label(item.get("quality") or ""),
                "season": season,
                "episode": episode,
                "size": int(item.get("size") or 0),
                "size_label": format_size(int(item.get("size") or 0)),
                "source_label": item.get("source_label") or "",
            }
        )
    panel_key = _panel_key_for_row(row)
    skip_reason = str(getattr(row, "skip_reason", "") or "")
    existing_content = dict(getattr(row, "existing_content", {}) or {})
    return {
        "id": str(row.id),
        "title": row.title,
        "content_type": row.content_type,
        "year": row.year or "",
        "panel": panel_key,
        "panel_label": _panel_label(panel_key),
        "tmdb_status": row.tmdb_status,
        "file_status": row.file_status,
        "upload_ready": bool(row.upload_ready),
        "uploaded": bool(row.uploaded),
        "uploaded_at": row.uploaded_at.isoformat() if row.uploaded_at else "",
        "poster_url": row.poster_url or "",
        "release_date": row.release_date or "",
        "missing_count": len(row.missing_items or []),
        "missing_items": list(row.missing_items or []),
        "live_notes": list(row.live_notes or []),
        "matched_files_count": len(row.matched_files or []),
        "matched_files_preview": matched_preview,
        "matched_files": matched_full,
        "season_quality_coverage": _series_quality_coverage(row),
        "file_choice_groups": list(getattr(row, "file_choice_groups", []) or []),
        "included_file_ids_count": len([str(x).strip() for x in (getattr(row, "included_file_ids", []) or []) if str(x).strip()]),
        "source_inputs": list(row.source_inputs or []),
        "upload_state": str(getattr(row, "upload_state", "") or "idle"),
        "upload_message": str(getattr(row, "upload_message", "") or ""),
        "last_error": row.last_error or "",
        "skip_reason": skip_reason,
        "existing_content": existing_content,
        "updated_at": row.updated_at.isoformat() if row.updated_at else "",
    }


def _panel_rank(panel: str) -> int:
    order = {
        "uploaded": 7,
        "uploading": 6,
        "complete": 5,
        "incomplete": 4,
        "file_not_found": 3,
        "tmdb_not_found": 2,
        "processing": 1,
        "skipped": 0,
    }
    key = _normalize_mass_panel_key(panel, fallback="")
    return order.get(key, 0)


def _row_identity_key(row: MassContentState) -> str:
    panel = (getattr(row, "panel", "") or "").strip().lower()
    tmdb_id = int(row.tmdb_id or 0)
    ctype = (row.content_type or "movie").strip().lower()
    if panel == "skipped":
        sig = _title_signature(row.title or "")
        if sig:
            return f"{ctype}:skipped:{sig}"
        return f"{ctype}:skipped:id:{row.id}"
    if tmdb_id > 0:
        return f"{ctype}:tmdb:{tmdb_id}"
    sig = _title_signature(row.title or "")
    if sig:
        return f"{ctype}:sig:{sig}"
    return f"{ctype}:id:{row.id}"


def _is_better_row(new_row: Any, current_row: Any) -> bool:
    new_tuple = (
        1 if new_row.uploaded else 0,
        1 if new_row.upload_ready else 0,
        _panel_rank(_panel_key_for_row(new_row)),
        new_row.updated_at or datetime.min,
    )
    cur_tuple = (
        1 if current_row.uploaded else 0,
        1 if current_row.upload_ready else 0,
        _panel_rank(_panel_key_for_row(current_row)),
        current_row.updated_at or datetime.min,
    )
    return new_tuple > cur_tuple


def _panel_key_for_row(row: Any) -> str:
    if bool(getattr(row, "uploaded", False)):
        return "uploaded"
    row_id = str(getattr(row, "id", "") or "")
    upload_state = str(getattr(row, "upload_state", "") or "idle").strip().lower()
    if upload_state in {"queued", "uploading"}:
        task = _MASS_UPLOAD_TASKS.get(row_id)
        if task and not task.done():
            return "uploading"
        # Stale/non-running upload state should fall back to actual panel.
        fallback_panel = _normalize_mass_panel_key(
            getattr(row, "panel", ""),
            fallback=("complete" if bool(getattr(row, "upload_ready", False)) else "incomplete"),
        )
        if fallback_panel == "uploading":
            fallback_panel = "complete" if bool(getattr(row, "upload_ready", False)) else "incomplete"
        return fallback_panel
    return _normalize_mass_panel_key(getattr(row, "panel", ""), fallback="processing")


def _dedupe_mass_rows(rows: list[Any]) -> list[Any]:
    picked: dict[str, Any] = {}
    for row in rows:
        identity = _row_identity_key(row)
        existing = picked.get(identity)
        if not existing or _is_better_row(row, existing):
            picked[identity] = row
    deduped = list(picked.values())
    deduped.sort(key=lambda x: x.updated_at or datetime.min, reverse=True)
    return deduped


async def _build_snapshot() -> dict:
    projection = {
        "title": 1,
        "content_type": 1,
        "year": 1,
        "panel": 1,
        "tmdb_status": 1,
        "file_status": 1,
        "upload_ready": 1,
        "uploaded": 1,
        "uploaded_at": 1,
        "poster_url": 1,
        "release_date": 1,
        "missing_items": {"$slice": [{"$ifNull": ["$missing_items", []]}, 300]},
        "live_notes": {"$slice": [{"$ifNull": ["$live_notes", []]}, 400]},
        "matched_files": {"$slice": [{"$ifNull": ["$matched_files", []]}, 220]},
        "seasons": {"$slice": [{"$ifNull": ["$seasons", []]}, 40]},
        "file_choice_groups": {"$slice": [{"$ifNull": ["$file_choice_groups", []]}, 160]},
        "included_file_ids": {"$slice": [{"$ifNull": ["$included_file_ids", []]}, 350]},
        "source_inputs": {"$slice": [{"$ifNull": ["$source_inputs", []]}, 50]},
        "upload_state": 1,
        "upload_message": 1,
        "last_error": 1,
        "skip_reason": 1,
        "existing_content": 1,
        "updated_at": 1,
        "tmdb_id": 1,
    }
    pipeline = [
        {"$sort": {"updated_at": -1}},
        {"$limit": _MASS_SNAPSHOT_ROW_LIMIT},
        {"$project": projection},
    ]
    raw_rows = await _collection_for(MassContentState).aggregate(pipeline).to_list(length=_MASS_SNAPSHOT_ROW_LIMIT)
    rows = [_MassRowProxy(doc) for doc in raw_rows]
    rows = _dedupe_mass_rows(rows)
    payload = [_serialize_row(row) for row in rows]
    panels = {
        "processing": [],
        "tmdb_not_found": [],
        "file_not_found": [],
        "incomplete": [],
        "complete": [],
        "uploading": [],
        "uploaded": [],
        "skipped": [],
    }
    for row in payload:
        panel = row.get("panel") or "processing"
        if panel not in panels:
            panel = "processing"
        panels[panel].append(row)
    return {
        "panels": panels,
        "counts": {key: len(value) for key, value in panels.items()},
        "total": len(payload),
        "import": _import_status_snapshot(),
        "workers": _mass_workers_snapshot(),
        "server_time": datetime.now().isoformat(),
    }


def _mass_workers_snapshot() -> dict[str, int]:
    return {
        "process_active": sum(1 for task in _MASS_TASKS.values() if not task.done()),
        "upload_active": sum(1 for task in _MASS_UPLOAD_TASKS.values() if not task.done()),
    }


def _mass_runtime_is_busy() -> bool:
    workers = _mass_workers_snapshot()
    if workers["process_active"] > 0 or workers["upload_active"] > 0:
        return True
    return bool(_MASS_IMPORT_STATUS.get("running"))


async def _build_snapshot_cached(force: bool = False) -> dict:
    now = asyncio.get_event_loop().time()
    cached = _MASS_SNAPSHOT_CACHE.get("data")
    last_ts = float(_MASS_SNAPSHOT_CACHE.get("ts") or 0.0)
    dirty = bool(_MASS_SNAPSHOT_CACHE.get("dirty"))
    if not force and cached is not None:
        age = now - last_ts
        if (not dirty) and (not _mass_runtime_is_busy()) and age <= _MASS_SNAPSHOT_IDLE_CACHE_TTL_SEC:
            return cached
        if age <= _MASS_SNAPSHOT_MIN_INTERVAL_SEC:
            return cached
    try:
        snapshot = await _build_snapshot()
    except Exception as exc:
        logger.warning("Mass snapshot build failed, using cached snapshot if available: %s", exc)
        if isinstance(cached, dict):
            return cached
        return {
            "panels": {
                "processing": [],
                "tmdb_not_found": [],
                "file_not_found": [],
                "incomplete": [],
                "complete": [],
                "uploading": [],
                "uploaded": [],
                "skipped": [],
            },
            "counts": {
                "processing": 0,
                "tmdb_not_found": 0,
                "file_not_found": 0,
                "incomplete": 0,
                "complete": 0,
                "uploading": 0,
                "uploaded": 0,
                "skipped": 0,
            },
            "total": 0,
            "import": _import_status_snapshot(),
            "workers": _mass_workers_snapshot(),
            "server_time": datetime.now().isoformat(),
            "warning": "Snapshot temporarily unavailable.",
        }
    _MASS_SNAPSHOT_CACHE["data"] = snapshot
    _MASS_SNAPSHOT_CACHE["ts"] = now
    _MASS_SNAPSHOT_CACHE["dirty"] = False
    return snapshot


def _invalidate_mass_snapshot_cache() -> None:
    _MASS_SNAPSHOT_CACHE["dirty"] = True
    _MASS_SNAPSHOT_CACHE["ts"] = 0.0


async def _mass_broadcast_snapshot(force: bool = False) -> None:
    async with _MASS_BROADCAST_LOCK:
        _MASS_SNAPSHOT_CACHE["dirty"] = True
        snapshot = await _build_snapshot_cached(force=bool(force))
        if not _MASS_WS_CLIENTS:
            return
        dead: list[WebSocket] = []
        for ws in list(_MASS_WS_CLIENTS):
            try:
                await ws.send_json({"type": "snapshot", "data": snapshot})
            except Exception:
                dead.append(ws)
        for ws in dead:
            _MASS_WS_CLIENTS.discard(ws)


async def _ws_auth_admin(websocket: WebSocket) -> User | None:
    phone = websocket.cookies.get("user_phone")
    if not phone:
        return None
    user = await User.find_one(User.phone_number == phone)
    if not user:
        return None
    if not _is_admin(user):
        return None
    return user


@router.get("/advance-mass-content-adder")
async def advance_mass_content_adder_page(request: Request):
    user = await get_current_user(request)
    if not user:
        return RedirectResponse("/admin-login")
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")

    _schedule_mass_dedupe()
    _schedule_mass_upload_recovery()
    base_ctx = await _admin_context_base(user)
    initial_state = await _build_snapshot_cached(force=True)
    return templates.TemplateResponse(
        "advance_mass_content_adder.html",
        {
            "request": request,
            **base_ctx,
            "tmdb_configured": bool(getattr(settings, "TMDB_API_KEY", "")),
            "initial_state": initial_state,
        },
    )


@router.get("/advance-mass-content-adder/state")
async def advance_mass_content_adder_state(request: Request):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")
    _schedule_mass_upload_recovery()
    return await _build_snapshot_cached(force=False)


@router.websocket("/advance-mass-content-adder/ws")
async def advance_mass_content_adder_ws(websocket: WebSocket):
    user = await _ws_auth_admin(websocket)
    if not user:
        await websocket.close(code=4403)
        return
    await websocket.accept()
    _MASS_WS_CLIENTS.add(websocket)
    try:
        await websocket.send_json({"type": "snapshot", "data": await _build_snapshot_cached(force=False)})
        while True:
            text = await websocket.receive_text()
            cmd = (text or "").strip().lower()
            if cmd in {"refresh", "ping"}:
                await websocket.send_json({"type": "snapshot", "data": await _build_snapshot_cached(force=(cmd == "refresh"))})
    except WebSocketDisconnect:
        pass
    except Exception:
        pass
    finally:
        _MASS_WS_CLIENTS.discard(websocket)


@router.post("/advance-mass-content-adder/manual-add")
async def advance_mass_content_adder_manual_add(
    request: Request,
    title: str = Form(""),
    content_type: str = Form(""),
    year: str = Form(""),
    force_add_existing: str = Form("0"),
):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")

    name = (title or "").strip()
    if not name:
        return JSONResponse({"ok": False, "error": "Title is required."}, status_code=400)
    allow_existing = (force_add_existing or "").strip().lower() in {"1", "true", "yes", "on"}
    exists, existing_payload = await _content_exists_for_input(name, content_type, (year or "").strip())
    if exists and not allow_existing:
        await _upsert_skipped_entry(
            title=name,
            content_type=content_type,
            year=(year or "").strip(),
            reason="Already exists in content database (manual add blocked).",
            existing_content=existing_payload,
        )
        asyncio.create_task(_mass_broadcast_snapshot())
        return JSONResponse(
            {
                "ok": False,
                "needs_existing_confirm": True,
                "message": "Content already exists in the database.",
                "existing": existing_payload,
            },
            status_code=409,
        )
    try:
        row = await _upsert_mass_entry(name, content_type, (year or "").strip(), "manual")
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

    _schedule_process(str(row.id), mode="full")
    _schedule_mass_dedupe()
    asyncio.create_task(_mass_broadcast_snapshot())
    return {"ok": True, "id": str(row.id)}


@router.post("/advance-mass-content-adder/import")
async def advance_mass_content_adder_import(request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")

    filename = (file.filename or "").strip().lower()
    if not filename.endswith(".csv") and not filename.endswith(".xlsx"):
        return JSONResponse({"ok": False, "error": "Only .csv and .xlsx files are supported."}, status_code=400)
    payload = await file.read()
    if not payload:
        return JSONResponse({"ok": False, "error": "Uploaded file is empty."}, status_code=400)

    try:
        if filename.endswith(".csv"):
            rows = _parse_csv_payload(payload)
        else:
            rows = _parse_xlsx_payload(payload)
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

    rows = _dedupe_rows(rows)
    if not rows:
        return JSONResponse({"ok": False, "error": "No valid content rows found."}, status_code=400)

    existing_index = await _build_existing_content_index()
    filtered_rows: list[dict] = []
    skipped_existing = 0
    for row in rows:
        title = (row.get("title") or "").strip()
        ctype = (row.get("type") or "").strip()
        year = (row.get("year") or "").strip()
        if _row_exists_in_content_index(title, ctype, year, existing_index):
            skipped_existing += 1
            existing_payload = _existing_detail_from_index(title, ctype, existing_index)
            await _upsert_skipped_entry(
                title=title,
                content_type=ctype,
                year=year,
                reason="Already exists in content database (matched by title/type; year ignored).",
                existing_content=existing_payload,
            )
            continue
        filtered_rows.append(row)

    rows = filtered_rows
    if not rows:
        return JSONResponse(
            {
                "ok": False,
                "error": "All imported entries already exist in content database.",
                "skipped_existing": skipped_existing,
            },
            status_code=409,
        )

    queued_count = await _enqueue_import_rows(rows)
    _schedule_mass_dedupe()
    asyncio.create_task(_mass_broadcast_snapshot())
    return {
        "ok": True,
        "count": len(rows),
        "queued": queued_count,
        "skipped_existing": skipped_existing,
        "chunk_size": _IMPORT_CHUNK_SIZE,
        "message": f"Queued {queued_count} rows for background processing in chunks of {_IMPORT_CHUNK_SIZE}."
                   + (f" Skipped {skipped_existing} existing entries." if skipped_existing else ""),
    }


@router.post("/advance-mass-content-adder/retry-tmdb/{item_id}")
async def advance_mass_content_adder_retry_tmdb(request: Request, item_id: str):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")
    row = await MassContentState.get(item_id)
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")
    row.panel = "processing"
    row.tmdb_status = "pending"
    row.file_status = "pending"
    row.uploaded = False
    row.uploaded_at = None
    row.upload_state = "idle"
    row.upload_message = ""
    row.last_error = None
    row.updated_at = datetime.now()
    await row.save()
    _schedule_process(str(row.id), mode="full")
    await _mass_broadcast_snapshot()
    return {"ok": True}


@router.post("/advance-mass-content-adder/retry-files/{item_id}")
async def advance_mass_content_adder_retry_files(request: Request, item_id: str):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")
    row = await MassContentState.get(item_id)
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")
    if row.tmdb_status != "found":
        return JSONResponse({"ok": False, "error": "TMDB data is missing for this item."}, status_code=400)
    row.panel = "processing"
    row.file_status = "pending"
    row.uploaded = False
    row.uploaded_at = None
    row.upload_state = "idle"
    row.upload_message = ""
    row.last_error = None
    row.updated_at = datetime.now()
    await row.save()
    _STORAGE_POOL_CACHE["rows"] = []
    _STORAGE_POOL_CACHE["expires_at"] = datetime.min
    _schedule_process(str(row.id), mode="files")
    await _mass_broadcast_snapshot()
    return {"ok": True}


@router.post("/advance-mass-content-adder/process-skipped/{item_id}")
async def advance_mass_content_adder_process_skipped(request: Request, item_id: str):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")
    row = await MassContentState.get(item_id)
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")

    row.panel = "processing"
    row.tmdb_status = "pending"
    row.file_status = "pending"
    row.upload_ready = False
    row.uploaded = False
    row.uploaded_at = None
    row.upload_state = "idle"
    row.upload_message = ""
    row.skip_reason = ""
    row.existing_content = {}
    row.last_error = None
    row.updated_at = datetime.now()
    await row.save()
    _schedule_process(str(row.id), mode="full")
    await _mass_broadcast_snapshot()
    return {"ok": True}


@router.get("/advance-mass-content-adder/storage/search")
async def advance_mass_content_storage_search(
    request: Request,
    q: str = "",
    offset: int = 0,
    limit: int = 60,
):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")

    q = (q or "").strip()
    try:
        offset = max(int(offset), 0)
    except Exception:
        offset = 0
    try:
        limit = int(limit)
    except Exception:
        limit = 60
    limit = max(20, min(limit, 200))

    query: dict[str, Any] = {
        "is_folder": False,
        "$or": [
            {"source": "bot"},
            {"source": "storage", "catalog_status": {"$nin": ["published", "used"]}},
            {"source": "upload", "catalog_status": {"$nin": ["published", "used"]}},
        ],
    }
    if q:
        search_regex = _build_title_regex(q)
        if not search_regex:
            return {"ok": True, "items": [], "has_more": False}
        query["name"] = {"$regex": search_regex, "$options": "i"}

    sort_field = "name" if q else "-created_at"
    rows = await FileSystemItem.find(query).sort(sort_field).skip(offset).limit(limit + 1).to_list()
    has_more = len(rows) > limit
    if has_more:
        rows = rows[:limit]

    payload = []
    for row in rows:
        if not _is_video_row(row):
            continue
        season_guess, episode_guess = _series_season_episode(row.name or "")
        quality_guess = _movie_quality_from_name_or_size(row.name or "", int(row.size or 0))
        payload.append(
            {
                "id": str(row.id),
                "name": row.name or "",
                "size": int(row.size or 0),
                "size_label": format_size(int(row.size or 0)),
                "season_guess": int(season_guess) if season_guess else None,
                "episode_guess": int(episode_guess) if episode_guess else None,
                "quality_guess": quality_guess,
            }
        )
    return {"ok": True, "items": payload, "has_more": has_more}


@router.post("/advance-mass-content-adder/attach-file/{item_id}")
async def advance_mass_content_adder_attach_file(
    request: Request,
    item_id: str,
    file_id: str = Form(""),
    quality: str = Form(""),
    season: str = Form(""),
    episode: str = Form(""),
):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")
    row = await MassContentState.get(item_id)
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")
    if row.tmdb_status != "found":
        return JSONResponse({"ok": False, "error": "TMDB data is missing for this item."}, status_code=400)

    file_id = (file_id or "").strip()
    if not file_id:
        return JSONResponse({"ok": False, "error": "file_id is required."}, status_code=400)
    storage_row = await FileSystemItem.get(file_id)
    if not storage_row or storage_row.is_folder:
        return JSONResponse({"ok": False, "error": "File not found."}, status_code=404)
    source = (storage_row.source or "").strip().lower()
    if source not in {"storage", "bot", "upload"}:
        return JSONResponse({"ok": False, "error": "Only bot/storage/upload files can be attached here."}, status_code=400)
    status = (storage_row.catalog_status or "").strip().lower()
    if status in {"published", "used"}:
        return JSONResponse({"ok": False, "error": "This file is already published/used."}, status_code=400)
    if not _is_video_row(storage_row):
        return JSONResponse({"ok": False, "error": "Selected item is not a video file."}, status_code=400)

    included = [str(x).strip() for x in (getattr(row, "included_file_ids", []) or []) if str(x).strip()]
    if file_id not in included:
        included.append(file_id)
    row.included_file_ids = included

    excluded = [str(x).strip() for x in (getattr(row, "excluded_file_ids", []) or []) if str(x).strip()]
    if file_id in excluded:
        excluded = [x for x in excluded if x != file_id]
    row.excluded_file_ids = excluded

    raw_quality = (quality or "").strip()
    if raw_quality:
        clean_quality = _normalize_quality_label(raw_quality)
        quality_map = dict(getattr(row, "file_quality_overrides", {}) or {})
        quality_map[file_id] = clean_quality
        row.file_quality_overrides = quality_map

    if row.content_type == "series":
        season_no = _int_or_none(season)
        episode_no = _int_or_none(episode)
        if season_no and episode_no:
            se_map = dict(getattr(row, "file_season_episode_overrides", {}) or {})
            se_map[file_id] = {"season": int(season_no), "episode": int(episode_no)}
            row.file_season_episode_overrides = se_map

    seed_rows = list(row.matched_files or [])
    guessed_quality = _normalize_quality_label(raw_quality) if raw_quality else (
        _series_quality_from_name(storage_row.name or "")
        if row.content_type == "series"
        else _movie_quality_from_name_or_size(storage_row.name or "", int(storage_row.size or 0))
    )
    season_guess, episode_guess = _series_season_episode(storage_row.name or "")
    seed_rows.append(
        {
            "file_id": file_id,
            "name": storage_row.name or "",
            "size": int(storage_row.size or 0),
            "quality": guessed_quality,
            "season": int(season_no) if row.content_type == "series" and season_no else (int(season_guess) if row.content_type == "series" and season_guess else None),
            "episode": int(episode_no) if row.content_type == "series" and episode_no else (int(episode_guess) if row.content_type == "series" and episode_guess else None),
            "source": (storage_row.source or "").strip().lower(),
            "source_label": _source_label(storage_row),
            "upload_label": _source_upload_label(storage_row),
        }
    )
    await _recompute_mass_item_files_fast(row, base_found_rows=seed_rows)
    return {"ok": True}


@router.post("/advance-mass-content-adder/reassign-quality/{item_id}")
async def advance_mass_content_adder_reassign_quality(
    request: Request,
    item_id: str,
    file_id: str = Form(""),
    quality: str = Form(""),
):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")
    row = await MassContentState.get(item_id)
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")
    if row.tmdb_status != "found":
        return JSONResponse({"ok": False, "error": "TMDB data is missing for this item."}, status_code=400)

    file_id = (file_id or "").strip()
    if not file_id:
        return JSONResponse({"ok": False, "error": "file_id is required."}, status_code=400)
    raw_quality = (quality or "").strip()
    if not raw_quality:
        return JSONResponse({"ok": False, "error": "quality is required."}, status_code=400)
    quality = _normalize_quality_label(raw_quality)

    matched_ids = {str(x.get("file_id") or "").strip() for x in (row.matched_files or [])}
    included_ids = {str(x).strip() for x in (getattr(row, "included_file_ids", []) or []) if str(x).strip()}
    if file_id not in matched_ids and file_id not in included_ids:
        return JSONResponse({"ok": False, "error": "File is not attached to this content."}, status_code=400)

    quality_map = dict(getattr(row, "file_quality_overrides", {}) or {})
    quality_map[file_id] = quality
    row.file_quality_overrides = quality_map

    selected_map = dict(getattr(row, "selected_file_map", {}) or {})
    for key, value in list(selected_map.items()):
        if str(value or "").strip() == file_id:
            selected_map.pop(key, None)
    row.selected_file_map = selected_map

    await _recompute_mass_item_files_fast(row, base_found_rows=list(row.matched_files or []))
    return {"ok": True}


@router.post("/advance-mass-content-adder/select-file/{item_id}")
async def advance_mass_content_adder_select_file(
    request: Request,
    item_id: str,
    bucket_key: str = Form(""),
    file_id: str = Form(""),
):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")
    row = await MassContentState.get(item_id)
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")
    if row.tmdb_status != "found":
        return JSONResponse({"ok": False, "error": "TMDB data is missing for this item."}, status_code=400)

    bucket_key = (bucket_key or "").strip()
    file_id = (file_id or "").strip()
    if not bucket_key or not file_id:
        return JSONResponse({"ok": False, "error": "bucket_key and file_id are required."}, status_code=400)

    matched_ids = {str(x.get("file_id") or "").strip() for x in (row.matched_files or [])}
    if file_id not in matched_ids:
        return JSONResponse({"ok": False, "error": "Selected file is not available in current matched list."}, status_code=400)

    selected_map = dict(getattr(row, "selected_file_map", {}) or {})
    selected_map[bucket_key] = file_id
    row.selected_file_map = selected_map
    await _recompute_mass_item_files_fast(row, base_found_rows=list(row.matched_files or []))
    return {"ok": True}


@router.post("/advance-mass-content-adder/remove-matched-file/{item_id}")
async def advance_mass_content_adder_remove_matched_file(
    request: Request,
    item_id: str,
    file_id: str = Form(""),
):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")
    row = await MassContentState.get(item_id)
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")

    file_id = (file_id or "").strip()
    if not file_id:
        return JSONResponse({"ok": False, "error": "file_id is required."}, status_code=400)

    matched_ids = {str(x.get("file_id") or "").strip() for x in (row.matched_files or [])}
    included_ids = {str(x).strip() for x in (getattr(row, "included_file_ids", []) or []) if str(x).strip()}
    if file_id and file_id not in matched_ids and file_id not in included_ids:
        return JSONResponse({"ok": False, "error": "File not found in matched list."}, status_code=400)

    excluded = [str(x).strip() for x in (getattr(row, "excluded_file_ids", []) or []) if str(x).strip()]
    if file_id not in excluded:
        excluded.append(file_id)
    row.excluded_file_ids = excluded

    included = [str(x).strip() for x in (getattr(row, "included_file_ids", []) or []) if str(x).strip()]
    if included:
        row.included_file_ids = [x for x in included if x != file_id]

    quality_map = dict(getattr(row, "file_quality_overrides", {}) or {})
    quality_map.pop(file_id, None)
    row.file_quality_overrides = quality_map

    se_map = dict(getattr(row, "file_season_episode_overrides", {}) or {})
    se_map.pop(file_id, None)
    row.file_season_episode_overrides = se_map

    selected_map = dict(getattr(row, "selected_file_map", {}) or {})
    for key, value in list(selected_map.items()):
        if str(value or "").strip() == file_id:
            selected_map.pop(key, None)
    row.selected_file_map = selected_map

    await _recompute_mass_item_files_fast(row, base_found_rows=list(row.matched_files or []))
    return {"ok": True}


@router.post("/advance-mass-content-adder/manual-file-hint/{item_id}")
async def advance_mass_content_adder_manual_file_hint(
    request: Request,
    item_id: str,
    hint: str = Form(""),
):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")
    row = await MassContentState.get(item_id)
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")
    if row.tmdb_status != "found":
        return JSONResponse({"ok": False, "error": "TMDB data is missing for this item."}, status_code=400)

    raw_hint = (hint or "").strip()
    if not raw_hint:
        return JSONResponse({"ok": False, "error": "Manual file hint is required."}, status_code=400)

    # Accept comma/newline separated tokens and merge as extra title markers for file scan.
    hints = [x.strip() for x in re.split(r"[\r\n,]+", raw_hint) if x.strip()]
    merged_inputs = list(row.source_inputs or [])
    added = 0
    for value in hints:
        marker = _to_title_marker(value)
        if marker and marker not in merged_inputs:
            merged_inputs.append(marker)
            added += 1
    if added > 0:
        row.source_inputs = merged_inputs

    row.panel = "processing"
    row.file_status = "pending"
    row.uploaded = False
    row.uploaded_at = None
    row.upload_state = "idle"
    row.upload_message = ""
    row.last_error = None
    row.updated_at = datetime.now()
    await row.save()
    _STORAGE_POOL_CACHE["rows"] = []
    _STORAGE_POOL_CACHE["expires_at"] = datetime.min
    _schedule_process(str(row.id), mode="files")
    await _mass_broadcast_snapshot()
    return {"ok": True, "added_hints": added}


@router.post("/advance-mass-content-adder/delete/{item_id}")
async def advance_mass_content_adder_delete(request: Request, item_id: str):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")
    row = await MassContentState.get(item_id)
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")
    await row.delete()
    await _mass_broadcast_snapshot()
    return {"ok": True}


_CLEARABLE_MASS_PANELS = set(_MASS_PANEL_KEYS)


def _effective_mass_panel(row: MassContentState) -> str:
    return _panel_key_for_row(row)


async def _cancel_worker_tasks_for_ids(item_ids: list[str]) -> dict[str, int]:
    keys = {str(x).strip() for x in (item_ids or []) if str(x).strip()}
    if not keys:
        return {"process": 0, "upload": 0}

    proc_tasks: list[asyncio.Task] = []
    upload_tasks: list[asyncio.Task] = []

    for key in keys:
        task = _MASS_TASKS.get(key)
        if task and not task.done():
            task.cancel()
            proc_tasks.append(task)
        task = _MASS_UPLOAD_TASKS.get(key)
        if task and not task.done():
            task.cancel()
            upload_tasks.append(task)

    if proc_tasks:
        try:
            await asyncio.wait_for(asyncio.gather(*proc_tasks, return_exceptions=True), timeout=1.5)
        except Exception:
            pass
    if upload_tasks:
        try:
            await asyncio.wait_for(asyncio.gather(*upload_tasks, return_exceptions=True), timeout=1.5)
        except Exception:
            pass

    for key in keys:
        if key in _MASS_TASKS and _MASS_TASKS[key].done():
            _MASS_TASKS.pop(key, None)
        if key in _MASS_UPLOAD_TASKS and _MASS_UPLOAD_TASKS[key].done():
            _MASS_UPLOAD_TASKS.pop(key, None)
        _MASS_LOCKS.pop(key, None)
        _MASS_LOCKS.pop(f"upload:{key}", None)

    return {"process": len(proc_tasks), "upload": len(upload_tasks)}


async def _cancel_import_worker(reason: str = "Cleared by admin.") -> bool:
    global _MASS_IMPORT_TASK
    task: asyncio.Task | None = None
    async with _MASS_IMPORT_LOCK:
        task = _MASS_IMPORT_TASK
        _MASS_IMPORT_QUEUE.clear()
        _MASS_IMPORT_STATUS["queued"] = 0
        _MASS_IMPORT_STATUS["running"] = False
        _MASS_IMPORT_STATUS["message"] = reason
        _MASS_IMPORT_STATUS["updated_at"] = _now_iso()

    was_running = bool(task and not task.done())
    if was_running and task is not None:
        task.cancel()
        try:
            await task
        except BaseException:
            pass

    async with _MASS_IMPORT_LOCK:
        _MASS_IMPORT_TASK = None
        _MASS_IMPORT_STATUS["queued"] = 0
        _MASS_IMPORT_STATUS["running"] = False
        _MASS_IMPORT_STATUS["message"] = reason
        _MASS_IMPORT_STATUS["updated_at"] = _now_iso()

    return was_running


@router.post("/advance-mass-content-adder/clear-panel/{panel}")
async def advance_mass_content_adder_clear_panel(request: Request, panel: str):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")

    key = _normalize_mass_panel_key(panel, fallback="")
    if key not in _CLEARABLE_MASS_PANELS:
        return JSONResponse({"ok": False, "error": "Invalid panel key."}, status_code=400)

    # Fast server-side filter instead of loading all rows in memory.
    if key == "uploaded":
        panel_filter: dict[str, Any] = {"uploaded": True}
    elif key == "uploading":
        panel_filter = {"uploaded": {"$ne": True}, "upload_state": {"$in": ["queued", "uploading"]}}
    else:
        panel_filter = {
            "uploaded": {"$ne": True},
            "upload_state": {"$nin": ["queued", "uploading"]},
            "panel": key,
        }

    deleted_count = 0
    target_id_strs: list[str] = []
    cursor = _collection_for(MassContentState).find(panel_filter, {"_id": 1}).limit(6000)
    async for row in cursor:
        raw_id = row.get("_id")
        if raw_id is not None:
            target_id_strs.append(str(raw_id))
    if not target_id_strs:
        return {"ok": True, "panel": key, "deleted": 0, "cancelled_process": 0, "cancelled_upload": 0}

    cancelled = await _cancel_worker_tasks_for_ids(target_id_strs)
    try:
        result = await _collection_for(MassContentState).delete_many(panel_filter)
        deleted_count = int(getattr(result, "deleted_count", 0) or 0)
    except Exception:
        # Fallback path: delete row-by-row if bulk delete fails.
        for raw_id in target_id_strs:
            try:
                row = await MassContentState.get(raw_id)
                if row:
                    await row.delete()
                    deleted_count += 1
            except Exception:
                continue
    _invalidate_mass_snapshot_cache()
    await _mass_broadcast_snapshot(force=True)
    return {
        "ok": True,
        "panel": key,
        "deleted": deleted_count or len(target_id_strs),
        "cancelled_process": int(cancelled.get("process") or 0),
        "cancelled_upload": int(cancelled.get("upload") or 0),
    }


@router.post("/advance-mass-content-adder/clear-all")
async def advance_mass_content_adder_clear_all(request: Request):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")

    proc_cancel = 0
    upload_cancel = 0
    if _MASS_TASKS:
        proc_tasks = [task for task in _MASS_TASKS.values() if task and not task.done()]
        proc_cancel = len(proc_tasks)
        for task in proc_tasks:
            task.cancel()
        if proc_tasks:
            await asyncio.gather(*proc_tasks, return_exceptions=True)
    if _MASS_UPLOAD_TASKS:
        upload_tasks = [task for task in _MASS_UPLOAD_TASKS.values() if task and not task.done()]
        upload_cancel = len(upload_tasks)
        for task in upload_tasks:
            task.cancel()
        if upload_tasks:
            await asyncio.gather(*upload_tasks, return_exceptions=True)
    _MASS_TASKS.clear()
    _MASS_UPLOAD_TASKS.clear()
    _MASS_LOCKS.clear()
    await _cancel_import_worker(reason="Cleared by admin.")

    result = await _collection_for(MassContentState).delete_many({})
    deleted = int(getattr(result, "deleted_count", 0) or 0)
    _invalidate_mass_snapshot_cache()
    await _mass_broadcast_snapshot(force=True)
    return {
        "ok": True,
        "deleted": deleted,
        "cancelled_process": proc_cancel,
        "cancelled_upload": upload_cancel,
    }


def _build_upload_payload(row: MassContentState) -> tuple[list[str], dict[str, dict], str]:
    plan = list(row.upload_plan or [])
    if not plan:
        return [], {}, "No files available for upload."

    raw_file_ids: list[str] = []
    overrides: dict[str, dict] = {}
    for entry in plan:
        file_id = (entry.get("file_id") or "").strip()
        if not file_id:
            continue
        raw_file_ids.append(file_id)
        payload = {"quality": (entry.get("quality") or "HD").upper()}
        if row.content_type == "series":
            payload["season"] = int(entry.get("season") or 1)
            payload["episode"] = int(entry.get("episode") or 1)
            payload["episode_title"] = (entry.get("episode_title") or "").strip()
        overrides[file_id] = payload

    raw_file_ids = list(dict.fromkeys(raw_file_ids))
    if not raw_file_ids:
        return [], {}, "No valid file ids found in upload plan."
    return raw_file_ids, overrides, ""


async def _content_exists_for_row(row: MassContentState) -> tuple[bool, dict]:
    ctype = (row.content_type or "movie").strip().lower()
    tmdb_id = int(row.tmdb_id or 0)
    title = (row.title or "").strip()

    doc = None
    if tmdb_id > 0:
        doc = await ContentItem.find_one(
            ContentItem.content_type == ctype,
            ContentItem.tmdb_id == tmdb_id,
            ContentItem.status == "published",
        )
    if not doc and title:
        query = {
            "content_type": ctype,
            "status": "published",
            "title": {"$regex": f"^{re.escape(title)}$", "$options": "i"},
        }
        doc = await ContentItem.find_one(query)
    if not doc and title:
        # Fallback by normalized title (year intentionally ignored).
        candidates = await ContentItem.find(
            ContentItem.content_type == ctype,
            ContentItem.status == "published",
        ).limit(2000).to_list()
        row_norm = _norm_title(title)
        for cand in candidates:
            if _norm_title(getattr(cand, "title", "") or "") != row_norm:
                continue
            doc = cand
            break

    if not doc:
        return False, {}

    payload = {
        "id": str(doc.id),
        "slug": (doc.slug or "").strip(),
        "title": doc.title or "",
        "year": doc.year or "",
        "content_type": doc.content_type or ctype,
    }
    return True, payload


async def _run_upload_worker(item_id: str, allow_incomplete: bool) -> None:
    lock = _MASS_LOCKS.setdefault(f"upload:{item_id}", asyncio.Lock())
    async with lock:
        async with _MASS_UPLOAD_SEMAPHORE:
            row = await MassContentState.get(item_id)
            if not row:
                return

            async def _set_upload_stage(message: str, *, force_snapshot: bool = False) -> None:
                row.upload_state = "uploading"
                row.upload_message = message
                row.updated_at = datetime.now()
                await row.save()
                await _mass_broadcast_snapshot(force=force_snapshot)

            row.upload_state = "uploading"
            row.upload_message = "Step 1/5: Preparing upload worker..."
            row.updated_at = datetime.now()
            await row.save()
            await _mass_broadcast_snapshot()

            try:
                if row.tmdb_status != "found":
                    row.upload_state = "failed"
                    row.upload_message = "TMDB data not found."
                    row.last_error = "TMDB data not found."
                    row.updated_at = datetime.now()
                    await row.save()
                    await _mass_broadcast_snapshot(force=True)
                    return
                if row.panel == "incomplete" and not allow_incomplete:
                    row.upload_state = "failed"
                    row.upload_message = "Incomplete content requires confirmation."
                    row.last_error = "Incomplete content requires confirmation."
                    row.updated_at = datetime.now()
                    await row.save()
                    await _mass_broadcast_snapshot(force=True)
                    return

                await _set_upload_stage("Step 2/5: Building upload plan...")
                raw_file_ids, overrides, build_err = _build_upload_payload(row)
                if build_err:
                    row.upload_state = "failed"
                    row.upload_message = build_err
                    row.last_error = build_err
                    row.updated_at = datetime.now()
                    await row.save()
                    await _mass_broadcast_snapshot(force=True)
                    return

                await _set_upload_stage("Step 3/5: Validating selected files...")
                items = await FileSystemItem.find(In(FileSystemItem.id, _cast_ids(raw_file_ids))).to_list()
                items = [item for item in items if not item.is_folder]
                if not items:
                    row.upload_state = "failed"
                    row.upload_message = "Selected files are not available anymore."
                    row.last_error = "Selected files are not available anymore."
                    row.updated_at = datetime.now()
                    await row.save()
                    await _mass_broadcast_snapshot(force=True)
                    return

                await _set_upload_stage("Step 4/5: Publishing content to database...")
                await _publish_items(
                    items=items,
                    catalog_type=row.content_type,
                    title=row.title,
                    year=(row.year or "").strip(),
                    desc=(row.description or "").strip(),
                    genres_list=list(row.genres or []),
                    actors_list=list(row.actors or []),
                    director=(row.director or "").strip(),
                    trailer_url=(row.trailer_url or "").strip(),
                    release_date=(row.release_date or "").strip(),
                    poster_url=(row.poster_url or "").strip(),
                    backdrop_url=(row.backdrop_url or "").strip(),
                    trailer_key=(row.trailer_key or "").strip(),
                    cast_profiles=list(row.cast_profiles or []),
                    tmdb_id=row.tmdb_id,
                    overrides=overrides,
                    sync_force=True,
                )
                _STORAGE_POOL_CACHE["rows"] = []
                _STORAGE_POOL_CACHE["expires_at"] = datetime.min

                await _set_upload_stage("Step 5/5: Finalizing catalog and cache...")
                row.uploaded = True
                row.uploaded_at = datetime.now()
                row.panel = "uploaded"
                row.file_status = "complete"
                row.upload_ready = True
                row.upload_state = "done"
                row.upload_message = "Upload completed successfully."
                row.last_error = None
                row.updated_at = datetime.now()
                await row.save()
                await _mass_broadcast_snapshot(force=True)
            except asyncio.CancelledError:
                row = await MassContentState.get(item_id)
                if row:
                    row.upload_state = "failed"
                    row.upload_message = "Upload worker interrupted. Click Upload again."
                    row.last_error = "Upload worker interrupted."
                    row.updated_at = datetime.now()
                    await row.save()
                    await _mass_broadcast_snapshot(force=True)
                raise
            except Exception as exc:
                row = await MassContentState.get(item_id)
                if row:
                    row.upload_state = "failed"
                    row.upload_message = str(exc)
                    row.last_error = str(exc)
                    row.updated_at = datetime.now()
                    await row.save()
                    await _mass_broadcast_snapshot(force=True)


def _queue_upload_worker(item_id: str, allow_incomplete: bool) -> bool:
    task = _MASS_UPLOAD_TASKS.get(item_id)
    if task and not task.done():
        return False

    async def _runner():
        try:
            await _run_upload_worker(item_id, allow_incomplete)
        finally:
            _MASS_UPLOAD_TASKS.pop(item_id, None)

    _MASS_UPLOAD_TASKS[item_id] = asyncio.create_task(_runner())
    return True


@router.post("/advance-mass-content-adder/upload/{item_id}")
async def advance_mass_content_adder_upload(
    request: Request,
    item_id: str,
    force_upload: str = Form("0"),
    force_reupload: str = Form("0"),
):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")

    row = await MassContentState.get(item_id)
    if not row:
        raise HTTPException(status_code=404, detail="Item not found")
    if row.tmdb_status != "found":
        return JSONResponse({"ok": False, "error": "TMDB data not found for this item."}, status_code=400)

    allow_incomplete = (force_upload or "").strip().lower() in {"1", "true", "yes", "on"}
    allow_reupload = (force_reupload or "").strip().lower() in {"1", "true", "yes", "on"}
    if row.panel == "incomplete" and not allow_incomplete:
        return JSONResponse(
            {
                "ok": False,
                "needs_confirm": True,
                "message": "Content is incomplete. Confirm to upload with missing files.",
                "missing_items": list(row.missing_items or []),
            },
            status_code=409,
        )

    exists, existing_payload = await _content_exists_for_row(row)
    if exists and not allow_reupload:
        return JSONResponse(
            {
                "ok": False,
                "needs_existing_confirm": True,
                "message": "Content already exists in content database. Do you want to upload again?",
                "existing": existing_payload,
            },
            status_code=409,
        )

    raw_ids, _, build_err = _build_upload_payload(row)
    if build_err:
        return JSONResponse({"ok": False, "error": build_err}, status_code=400)
    if not raw_ids:
        return JSONResponse({"ok": False, "error": "No files available for upload."}, status_code=400)

    queued = _queue_upload_worker(item_id, allow_incomplete)
    if not queued:
        row.upload_state = str(getattr(row, "upload_state", "") or "uploading")
        row.upload_message = "Upload already running for this content."
        row.updated_at = datetime.now()
        await row.save()
        await _mass_broadcast_snapshot()
        return JSONResponse(
            {
                "ok": True,
                "queued": False,
                "already_running": True,
                "message": "Upload already running for this content.",
            },
            status_code=200,
        )

    row.upload_state = "queued"
    row.upload_message = "Step 0/5: Queued. Waiting for worker slot..."
    row.updated_at = datetime.now()
    await row.save()
    await _mass_broadcast_snapshot()
    return JSONResponse(
        {
            "ok": True,
            "queued": True,
            "already_running": False,
            "message": "Upload queued in background.",
        },
        status_code=202,
    )


def _rows_for_panel(panel: str, rows: list[MassContentState]) -> list[MassContentState]:
    key = _normalize_mass_panel_key(panel, fallback="")
    valid = _MASS_PANEL_KEYS
    if key not in valid:
        return []
    out = []
    for row in rows:
        row_panel = _panel_key_for_row(row)
        if row_panel == key:
            out.append(row)
    return out


@router.get("/advance-mass-content-adder/export/{panel}")
async def advance_mass_content_adder_export(request: Request, panel: str):
    user = await get_current_user(request)
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail="Not authorized.")

    panel_key = _normalize_mass_panel_key(panel, fallback="")
    if panel_key not in _MASS_PANEL_KEYS:
        return JSONResponse({"ok": False, "error": "Invalid panel."}, status_code=400)

    query: dict[str, Any] = {}
    if panel_key == "uploaded":
        query["uploaded"] = True
    elif panel_key == "uploading":
        query["upload_state"] = {"$in": ["queued", "uploading"]}
    else:
        query["panel"] = panel_key

    rows = await MassContentState.find(query).sort("-updated_at").limit(_MASS_EXPORT_ROW_LIMIT).to_list()
    rows = _dedupe_mass_rows(rows)
    filtered = _rows_for_panel(panel_key, rows)

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(
        [
            "title",
            "type",
            "year",
            "panel",
            "tmdb_status",
            "file_status",
            "upload_ready",
            "uploaded",
            "missing_count",
            "missing_details",
            "skip_reason",
            "existing_content_id",
            "existing_content_title",
            "existing_content_type",
            "existing_content_year",
            "existing_content_slug",
        ]
    )
    for row in filtered:
        missing_details = " | ".join([(x.get("note") or "") for x in (row.missing_items or []) if (x.get("note") or "")])
        existing_payload = dict(getattr(row, "existing_content", {}) or {})
        writer.writerow(
            [
                row.title,
                row.content_type,
                row.year or "",
                _panel_key_for_row(row),
                row.tmdb_status,
                row.file_status,
                "yes" if row.upload_ready else "no",
                "yes" if row.uploaded else "no",
                len(row.missing_items or []),
                missing_details,
                (getattr(row, "skip_reason", "") or ""),
                existing_payload.get("id", ""),
                existing_payload.get("title", ""),
                existing_payload.get("content_type", ""),
                existing_payload.get("year", ""),
                existing_payload.get("slug", ""),
            ]
        )

    payload = out.getvalue().encode("utf-8")
    filename = f"mass_adder_{panel}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    headers = {"Content-Disposition": f'attachment; filename=\"{filename}\"'}
    return StreamingResponse(io.BytesIO(payload), media_type="text/csv; charset=utf-8", headers=headers)
